import streamlit as st
import pdfplumber
import gspread
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload
import pandas as pd
import re
from datetime import datetime, timedelta
import io
import json
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import difflib
import openpyxl
from openpyxl.styles import Border, Side, Alignment
import time
import random

# ==========================================
# 💎 PREMIUM ERP ARAYÜZ (UI/UX) CSS KODLARI
# ==========================================
st.set_page_config(page_title="Aktürk ERP v9.60", page_icon="🛡️", layout="wide", initial_sidebar_state="auto")

gizleme_kodu = """
<style>
#MainMenu {visibility: hidden;}
header {visibility: hidden;}    
footer {visibility: hidden;}    
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');
html, body, [class*="css"] { font-family: 'Inter', sans-serif !important; }
.stApp { background-color: #F0F2F5 !important; color: #1E293B !important; }
h1, h2, h3, h4 { color: #0F172A !important; font-weight: 800 !important; letter-spacing: -0.5px; }
input, textarea, .stTextInput input, .stTextArea textarea, .stNumberInput input { caret-color: #2563EB !important; color: #0F172A !important; font-weight: 600 !important; }
label, div[data-testid="stWidgetLabel"] > div > p { color: #0F172A !important; font-weight: 700 !important; font-size: 14px !important; }
.stRadio div[role="radiogroup"] label p, .stCheckbox label p { color: #1E293B !important; font-weight: 600 !important; }
.stSelectbox>div>div>div { color: #0F172A !important; font-weight: 600 !important; }
p { color: #334155 !important; }
.stTabs [data-baseweb="tab-list"] { background-color: transparent; border-bottom: 2px solid #E2E8F0; gap: 20px; }
.stTabs [data-baseweb="tab"] { background-color: transparent !important; border: none !important; padding: 10px 4px !important; }
.stTabs [data-baseweb="tab"] p { color: #64748B !important; font-weight: 700 !important; font-size: 14px !important; }
.stTabs [aria-selected="true"] p { color: #1D4ED8 !important; font-weight: 800 !important; }
.stTabs [aria-selected="true"] { border-bottom: 3px solid #2563EB !important; }
[data-testid="stSidebar"] { background-color: #FFFFFF !important; border-right: 1px solid #E2E8F0 !important; box-shadow: 2px 0 10px rgba(0,0,0,0.02); }
[data-testid="stSidebar"] hr { border-color: #E2E8F0 !important; margin: 15px 0; }
[data-testid="stSidebar"] div[data-baseweb="radio"] > div:first-child { display: none !important; }
[data-testid="stSidebar"] .stRadio div[role="radiogroup"] label { padding: 12px 16px !important; margin-bottom: 6px !important; border-radius: 8px !important; transition: all 0.2s ease-in-out !important; cursor: pointer !important; width: 100% !important; border: 1px solid transparent; }
[data-testid="stSidebar"] .stRadio div[role="radiogroup"] label p { color: #475569 !important; margin: 0 !important; font-weight: 600 !important; font-size: 15px !important;}
[data-testid="stSidebar"] .stRadio div[role="radiogroup"] label:hover { background-color: #F8FAFC !important; border: 1px solid #E2E8F0; transform: translateX(4px); }
[data-testid="stSidebar"] .stRadio div[role="radiogroup"] label[data-checked="true"], [data-testid="stSidebar"] .stRadio div[role="radiogroup"] label:has(input:checked) { background: linear-gradient(90deg, #EFF6FF 0%, #FFFFFF 100%) !important; border-left: 4px solid #2563EB !important; border-top: 1px solid #E2E8F0; border-bottom: 1px solid #E2E8F0; border-right: 1px solid #E2E8F0; }
[data-testid="stSidebar"] .stRadio div[role="radiogroup"] label[data-checked="true"] p, [data-testid="stSidebar"] .stRadio div[role="radiogroup"] label:has(input:checked) p { color: #1D4ED8 !important; font-weight: 800 !important; }
div[data-testid="stMetric"] { background-color: #FFFFFF !important; border-radius: 12px !important; padding: 20px !important; border: 1px solid #E2E8F0 !important; border-left: 5px solid #2563EB !important; box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05), 0 2px 4px -1px rgba(0, 0, 0, 0.03) !important; transition: transform 0.2s ease; }
div[data-testid="stMetric"]:hover { transform: translateY(-3px); box-shadow: 0 10px 15px -3px rgba(0, 0, 0, 0.08); }
div[data-testid="stMetricValue"] { color: #0F172A !important; font-weight: 800 !important; font-size: 28px !important; }
div[data-testid="stMetricLabel"] { color: #64748B !important; font-size: 14px !important; font-weight: 700 !important; text-transform: uppercase; letter-spacing: 0.5px; }
div[data-baseweb="select"] > div, div[data-baseweb="input"] > div, textarea { background-color: #FFFFFF !important; border: 1px solid #CBD5E1 !important; border-radius: 8px !important; transition: all 0.2s ease; }
div[data-baseweb="select"] > div:hover, div[data-baseweb="input"] > div:hover { border-color: #94A3B8 !important; }
.stButton>button[kind="primary"] { background: linear-gradient(135deg, #1D4ED8 0%, #1E3A8A 100%) !important; color: #FFFFFF !important; font-weight: 600 !important; font-size: 15px !important; border-radius: 8px !important; border: none !important; padding: 10px 24px !important; box-shadow: 0 4px 12px rgba(29, 78, 216, 0.3) !important; transition: all 0.3s ease !important; }
.stButton>button[kind="primary"]:hover { transform: translateY(-2px); box-shadow: 0 8px 16px rgba(29, 78, 216, 0.4) !important; }
.stButton>button[kind="secondary"] { background-color: #FFFFFF !important; color: #334155 !important; font-weight: 600 !important; border-radius: 8px !important; border: 1px solid #CBD5E1 !important; padding: 10px 24px !important; box-shadow: 0 1px 2px rgba(0,0,0,0.05) !important; transition: all 0.2s ease !important; }
.stButton>button[kind="secondary"]:hover { background-color: #F8FAFC !important; border-color: #94A3B8 !important; }
div[data-testid="stForm"] { background-color: #FFFFFF; padding: 30px; border-radius: 12px; border: 1px solid #E2E8F0; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.02); }
.login-box { background-color: #FFFFFF; padding: 40px; border-radius: 16px; border: 1px solid #E2E8F0; text-align: center; max-width: 420px; margin: auto; margin-top: 8vh; box-shadow: 0 20px 25px -5px rgba(0, 0, 0, 0.1), 0 8px 10px -6px rgba(0, 0, 0, 0.05); }
[data-testid="stDataEditor"] textarea, [data-testid="stDataEditor"] input, .glide-data-grid textarea, .gdg-input { background-color: #FFFFFF !important; color: #1E293B !important; font-weight: 600 !important; }
</style>
"""
st.markdown(gizleme_kodu, unsafe_allow_html=True)

# ==========================================
# 1. TEMEL AYARLAR VE SABİTLER
# ==========================================
VERSIYON = "v9.60 (Merkez Kasa & Profesyonel Excel İhracı)"
SHEET_ID = "19zBeYZMLjpMe5rx1d6p6TNwQjHGFfqAx-qVKVxDxh24"
DRIVE_KLASOR_ID = "17wXJilHVDuHhDWS-POS4nr_RjUZnN7eL" 

try:
    PORTAL_KULLANICI = st.secrets["PORTAL_KULLANICI"]
    PORTAL_SIFRE = st.secrets["PORTAL_SIFRE"]
    GONDEREN_MAIL = st.secrets["GONDEREN_MAIL"]
    MAIL_SIFRE = st.secrets["MAIL_SIFRE"]
except Exception:
    st.error("🚨 Güvenlik Kasası Bulunamadı! Lütfen bilgisayarınızdaki .streamlit/secrets.toml dosyanızı oluşturun.")
    st.stop()

# ==========================================
# 2. YARDIMCI FONKSİYONLAR VE API KALKANI
# ==========================================
if "giris_yapildi" not in st.session_state: st.session_state["giris_yapildi"] = False

def ekran_temizle():
    for key in list(st.session_state.keys()):
        if key not in ["giris_yapildi", "kullanici_adi", "google_kasa"]: del st.session_state[key]

def api_kalkani(fonksiyon):
    bekleme_suresi = 2
    for deneme in range(4):
        try: return fonksiyon()
        except Exception as e:
            if deneme < 3: time.sleep(bekleme_suresi); bekleme_suresi *= 2 
            else: st.error("🚨 Google Sunucuları şu an çok yoğun olduğu için yanıt vermiyor. Lütfen sayfayı yenileyip 1 dakika sonra tekrar deneyin."); return None

@st.cache_resource(show_spinner="Bağlantı Kuruluyor...")
def get_credentials():
    try:
        if "google_kasa" in st.secrets:
            return Credentials.from_service_account_info(json.loads(st.secrets["google_kasa"]), scopes=["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"])
    except Exception: pass
    st.error("🚨 Google Kasa verisi okunamadı. secrets.toml dosyasını kontrol edin."); st.stop()

@st.cache_resource
def get_client(): return gspread.authorize(get_credentials())
def get_drive_service(): return build('drive', 'v3', credentials=get_credentials())
client = get_client()

def temiz_isim(metin):
    if pd.isna(metin) or not metin: return ""
    return str(metin).strip().replace('i', 'İ').replace('ı', 'I').replace('ğ', 'Ğ').replace('ü', 'Ü').replace('ş', 'Ş').replace('ö', 'Ö').replace('ç', 'Ç').upper()

def sayiya_cevir(deger):
    if pd.isna(deger) or str(deger).strip() == "": return 0.0
    if isinstance(deger, (int, float)): return float(deger)
    deger_str = re.sub(r'[^\d.,-]', '', str(deger).strip()).rstrip('.,')
    if not deger_str: return 0.0
    if '.' in deger_str and ',' in deger_str:
        if deger_str.rfind(',') > deger_str.rfind('.'): deger_str = deger_str.replace('.', '').replace(',', '.')
        else: deger_str = deger_str.replace(',', '')
    elif ',' in deger_str:
        if deger_str.count(',') > 1: deger_str = deger_str.replace(',', '')
        else: deger_str = deger_str.replace(',', '.')
    elif '.' in deger_str:
        parts = deger_str.split('.')
        if len(parts) >= 2:
            if len(parts[-1]) != 3:
                son_nokta = deger_str.rfind('.')
                deger_str = deger_str[:son_nokta].replace('.', '') + '.' + deger_str[son_nokta+1:]
            else: deger_str = deger_str.replace('.', '')
    try: return float(deger_str)
    except: return 0.0

def para_format(deger):
    try: return "{:,.2f}".format(sayiya_cevir(deger)).replace(",", "X").replace(".", ",").replace("X", ".") + " TL"
    except: return "0,00 TL"

def editor_icin_hazirla(deger):
    try:
        v = sayiya_cevir(deger)
        if v == 0: return "0"
        if v.is_integer(): return str(int(v))
        return f"{v:.2f}".replace('.', ',')
    except: return "0"

def df_gorsel_yap(df, para_sutunlari):
    df_gorsel = df.copy()
    for col in para_sutunlari:
        if col in df_gorsel.columns: df_gorsel[col] = df_gorsel[col].apply(para_format)
    if "PDF Linki" in df_gorsel.columns: df_gorsel["PDF Linki"] = df_gorsel["PDF Linki"].apply(lambda x: None if pd.isna(x) or str(x).strip() in ["Yok", ""] else x)
    return df_gorsel

def hesap_kodu_ekle(df, hesap_tipi):
    if df.empty: return df
    kolon = "Kisi_Kurum" if "Kisi_Kurum" in df.columns else ("Kişi / Kurum / Acente" if "Kişi / Kurum / Acente" in df.columns else None)
    if not kolon: return df
    
    hk_map = {}
    if hesap_tipi == "Müşteri":
        df_mus = get_data("Musteriler")
        if not df_mus.empty and "Hesap_Kodu" in df_mus.columns and "Musteri_Adi" in df_mus.columns: hk_map = dict(zip(df_mus["Musteri_Adi"], df_mus["Hesap_Kodu"]))
        prefix = "120"
    elif hesap_tipi == "Tali":
        df_acn = get_data("Ayarlar_Acenteler")
        if not df_acn.empty and "Hesap_Kodu" in df_acn.columns and "Acente_Adi" in df_acn.columns: hk_map = dict(zip(df_acn["Acente_Adi"], df_acn["Hesap_Kodu"]))
        prefix = "320.T"
    else: 
        df_sir = get_data("Ayarlar_Sirketler")
        if not df_sir.empty and "Hesap_Kodu" in df_sir.columns and "Sirket_Adi" in df_sir.columns: hk_map = dict(zip(df_sir["Sirket_Adi"], df_sir["Hesap_Kodu"]))
        prefix = "320.S"
        
    def kod_bul(isim):
        isim = str(isim).strip()
        if isim in hk_map and str(hk_map[isim]).strip(): return hk_map[isim]
        uniques = sorted([x for x in df[kolon].dropna().unique() if str(x).strip() != ""])
        idx = uniques.index(isim) if isim in uniques else 0
        return f"{prefix}.{str(idx+1).zfill(3)}"
        
    df.insert(0, "Hesap Kodu", df[kolon].apply(kod_bul))
    return df

def excel_indir(df, buton_metni, dosya_adi):
    df_export = df.copy()
    if not df_export.empty:
        toplamlar = {}
        satir_sayisi = len(df_export)
        para_kolonlari = ["Net Prim", "Brüt Prim", "Şirket Komisyonu", "Aktürk Sigorta Kazancı", "Borc", "Alacak", "Toplam Borç", "Toplam Alacak", "Dip Tutar (Bakiye)", "Bakiye (Kalan)", "Hakediş (Bize Borç)", "Tahsilat (Bize Ödenen)"]
        
        borc_col = next((c for c in ["Borc", "Hakediş (Bize Borç)", "Toplam Borç"] if c in df_export.columns), None)
        alacak_col = next((c for c in ["Alacak", "Tahsilat (Bize Ödenen)", "Toplam Alacak"] if c in df_export.columns), None)

        for i, col in enumerate(df_export.columns):
            harf = chr(65 + i) if i < 26 else chr(64 + i // 26) + chr(65 + i % 26)
            if col in para_kolonlari:
                if col in ["Dip Tutar (Bakiye)", "Bakiye (Kalan)"] and borc_col and alacak_col:
                    borc_harf = chr(65 + list(df_export.columns).index(borc_col)) if list(df_export.columns).index(borc_col) < 26 else chr(64 + list(df_export.columns).index(borc_col) // 26) + chr(65 + list(df_export.columns).index(borc_col) % 26)
                    alacak_harf = chr(65 + list(df_export.columns).index(alacak_col)) if list(df_export.columns).index(alacak_col) < 26 else chr(64 + list(df_export.columns).index(alacak_col) // 26) + chr(65 + list(df_export.columns).index(alacak_col) % 26)
                    toplamlar[col] = f"=SUM({borc_harf}2:{borc_harf}{satir_sayisi+1})-SUM({alacak_harf}2:{alacak_harf}{satir_sayisi+1})"
                else:
                    toplamlar[col] = f"=SUM({harf}2:{harf}{satir_sayisi+1})"
            elif col == df_export.columns[0]: toplamlar[col] = "GENEL TOPLAM"
            else: toplamlar[col] = ""
        
        df_export = pd.concat([df_export, pd.DataFrame([toplamlar])], ignore_index=True)
        
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine='openpyxl') as writer: 
        df_export.to_excel(writer, index=False, sheet_name='Rapor')
        worksheet = writer.sheets['Rapor']
        ince_kenarlik = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = ince_kenarlik
                cell.alignment = Alignment(vertical='center')
                if cell.row > 1 and cell.value is not None:
                    col_name = df_export.columns[cell.column - 1]
                    if col_name in para_kolonlari:
                        try:
                            if isinstance(cell.value, str) and not cell.value.startswith("="): cell.value = sayiya_cevir(cell.value)
                            cell.number_format = '#,##0.00' 
                        except: pass
    st.download_button(f"📥 {buton_metni}", out.getvalue(), f"{dosya_adi}.xlsx")

def tarih_formatla(tarih_degeri):
    if pd.isna(tarih_degeri) or str(tarih_degeri).strip() == "": return datetime.now().strftime("%d.%m.%Y")
    try: return pd.to_datetime(str(tarih_degeri).strip(), dayfirst=True).strftime("%d.%m.%Y")
    except: return datetime.now().strftime("%d.%m.%Y")

def get_kok_police(pno):
    if pd.isna(pno): return ""
    match = re.search(r'\(Asıl Poliçe No:\s*(.*?)\)', str(pno))
    if match: return match.group(1).strip()
    return str(pno).strip()

def fis_no_uret(islem_tipi="POL"): return f"{islem_tipi}-{datetime.now().strftime('%y%m%d%H%M')}{random.randint(10, 99)}"
    
@st.cache_data(ttl=5, show_spinner=False)
def get_data(sheet_name):
    def _fetch():
        try: return client.open_by_key(SHEET_ID).worksheet(sheet_name).get_all_values()
        except gspread.exceptions.WorksheetNotFound: return []
    raw_data = api_kalkani(_fetch)
    if not raw_data: return pd.DataFrame()
    headers = raw_data[0]
    max_len = len(headers)
    rows = []
    for r in raw_data[1:]:
        if len(r) < max_len: r = r + [""] * (max_len - len(r))
        elif len(r) > max_len: r = r[:max_len]
        rows.append(r)
    df = pd.DataFrame(rows, columns=headers)
    if not df.empty:
        df['Sheet_Row'] = df.index + 2 
        for col in ["Müşteri Adı Soyadı", "Kisi_Kurum", "Musteri_Adi", "Sirket_Adi", "Sigorta Şirketi", "Acente", "Acente_Adi"]:
            if col in df.columns: df[col] = df[col].apply(temiz_isim)
    return df

def drive_pdf_yukle(file_bytes, file_name):
    def _yukle():
        drive_service = get_drive_service()
        file_metadata = {'name': file_name, 'parents': [DRIVE_KLASOR_ID]}
        media = MediaIoBaseUpload(io.BytesIO(file_bytes), mimetype='application/pdf', resumable=True)
        file = drive_service.files().create(body=file_metadata, media_body=media, fields='id, webViewLink', supportsAllDrives=True).execute()
        drive_service.permissions().create(fileId=file.get('id'), body={'type': 'anyone', 'role': 'reader'}, supportsAllDrives=True).execute()
        return file.get('webViewLink')
    link = api_kalkani(_yukle)
    return link if link else "Yok"

def drive_pdf_sil(link):
    if pd.isna(link) or link == "Yok" or not link: return
    try:
        match = re.search(r'/d/([a-zA-Z0-9_-]+)', str(link))
        if match:
            def _sil(): get_drive_service().files().delete(fileId=match.group(1), supportsAllDrives=True).execute(); return True
            api_kalkani(_sil)
    except: pass 

def klasik_analiz(metin):
    data = {"tanzim": "", "baslangic": "", "bitis": "", "musteri": "", "tc_vkn": "", "sirket": "", "urun": "", "p_no": "", "plaka": "", "net_prim": 0.0, "brut_prim": 0.0}
    metin_upper = metin.upper()
    sirketler = {"ANKARA SİGORTA": "Ankara Sigorta", "DOĞA SİGORTA": "Doğa Sigorta", "ALLIANZ": "Allianz Sigorta", "HDI SİGORTA": "HDI Sigorta", "HDİ": "HDI Sigorta", "HEPİYİ": "Hepiyi Sigorta", "RAY SİGORTA": "Ray Sigorta", "SOMPO": "Sompo Sigorta", "TÜRKİYE SİGORTA": "Türkiye Sigorta", "AK SİGORTA": "Ak Sigorta", "ETHICA": "Ethica Sigorta"}
    for anahtar, deger in sirketler.items():
        if anahtar in metin_upper: data["sirket"] = deger; break
    urun_tipleri = {"TRAFİK": "Trafik Sigortası", "KASKO": "Kasko", "SAĞLIK": "Sağlık Sigortası", "TSS": "Sağlık Sigortası", "DASK": "Dask", "DOĞAL AFET": "Dask", "KONUT": "Konut Sigortası", "İŞYERİ": "İşyeri Sigortası", "İMM": "İmm", "ALLRİSK": "İnşaat Allrisk"}
    for anahtar, deger in urun_tipleri.items():
        if anahtar in metin_upper: data["urun"] = deger; break
    tarihler = re.findall(r'\b\d{2}[\./-]\d{2}[\./-]\d{4}\b', metin)
    if len(tarihler) >= 3: data["tanzim"], data["baslangic"], data["bitis"] = [t.replace("/", ".") for t in tarihler[:3]]
    tc = re.search(r'\b[0-9]{10,11}\b', metin)
    if tc: data["tc_vkn"] = tc.group()
    plaka = re.search(r'\b(?:[0-8][0-9]|9[0-8])\s*[A-Z]{1,3}\s*[0-9]{2,4}\b', metin_upper)
    if plaka: data["plaka"] = plaka.group().replace(" ", "")
    return data

STIL_AYARLARI = {"PDF Linki": st.column_config.LinkColumn("📄 Belge", display_text="📥 PDF'İ AÇ")}

# ==========================================
# GİRİŞ SİSTEMİ 
# ==========================================
if not st.session_state["giris_yapildi"]:
    st.markdown("<div class='login-box'>", unsafe_allow_html=True)
    st.markdown("""<div style="background: linear-gradient(135deg, #1D4ED8 0%, #1E3A8A 100%); width: 88px; height: 88px; border-radius: 20px; display: flex; align-items: center; justify-content: center; margin: 0 auto 20px auto; box-shadow: 0 10px 25px rgba(29, 78, 216, 0.4); transform: rotate(-5deg);"><span style="font-size: 48px; font-weight: 900; color: #FFFFFF; transform: rotate(5deg);">A</span></div><h2 style='margin-top:0px; margin-bottom: 5px; color:#0F172A; font-weight: 800;'>Aktürk Sigorta</h2><p style='color: #64748B; font-size: 15px; margin-bottom: 30px; font-weight: 600;'>Kurumsal Yönetim Portalı</p>""", unsafe_allow_html=True)
    tab1, tab2 = st.tabs(["🔐 Güvenli Giriş", "🔑 Şifre İşlemleri"])
    with tab1:
        u = st.text_input("Kullanıcı Adı")
        p = st.text_input("Şifre", type="password")
        if st.button("Sisteme Giriş Yap", type="primary", use_container_width=True):
            if u == PORTAL_KULLANICI and p == PORTAL_SIFRE: st.session_state["giris_yapildi"] = True; st.session_state["kullanici_adi"] = u; st.rerun()
            else:
                with st.spinner("Doğrulanıyor..."):
                    df_u = get_data("Kullanicilar")
                    giris_onay = False
                    if not df_u.empty:
                        k_kol = "Kullanıcı Adı" if "Kullanıcı Adı" in df_u.columns else "Kullanici_Adi"
                        s_kol = "Şifre" if "Şifre" in df_u.columns else "Sifre"
                        if k_kol in df_u.columns and s_kol in df_u.columns:
                            eslesen = df_u[df_u[k_kol] == u]
                            if not eslesen.empty and str(eslesen[s_kol].values[0]) == str(p): giris_onay = True
                if giris_onay: st.session_state["giris_yapildi"] = True; st.session_state["kullanici_adi"] = u; st.rerun()
                else: st.error("⚠️ Hatalı kullanıcı adı veya şifre!")
    st.markdown("</div>", unsafe_allow_html=True)

# ==========================================
# ANA UYGULAMA
# ==========================================
else:
    st.sidebar.markdown("""<div style='display: flex; align-items: center; margin-bottom: 30px; margin-top: 10px; padding: 10px; background-color: #F8FAFC; border-radius: 12px; border: 1px solid #E2E8F0;'><div style='background: linear-gradient(135deg, #1D4ED8 0%, #1E3A8A 100%); width: 42px; height: 42px; border-radius: 10px; display: flex; align-items: center; justify-content: center; margin-right: 15px; box-shadow: 0 4px 10px rgba(29, 78, 216, 0.3);'><span style='font-size: 20px; font-weight: 900; color: #FFFFFF;'>A</span></div><div><h2 style='margin: 0; color: #0F172A; font-size: 18px; font-weight: 800; letter-spacing: -0.5px;'>Aktürk ERP</h2><span style='color: #64748B; font-size: 12px; font-weight: 700;'>Kullanıcı: {0}</span></div></div>""".format(st.session_state['kullanici_adi'].upper()), unsafe_allow_html=True)
    menu = st.sidebar.radio("", ["📥 İşlem Merkezi (Poliçe)", "💰 Cari & Mutabakat", "📅 Yenileme Takvimi", "🔎 Akıllı Arama", "🛠️ Kayıt Onarım & Silme", "⚙️ Sistem Ayarları", "📂 Genel Arşiv"], on_change=ekran_temizle, label_visibility="collapsed")
    st.sidebar.markdown(f"<div style='text-align: center; margin-top: 20px;'><span style='color: #94A3B8; font-size: 11px; font-weight: 600;'>{VERSIYON}</span></div>", unsafe_allow_html=True)

    if menu == "📥 İşlem Merkezi (Poliçe)":
        st.header("📥 Yeni Poliçe & İşlem Kaydı")
        df_urun = get_data("Ayarlar_Urunler")
        urun_listesi = df_urun["Urun_Adi"].tolist() if not df_urun.empty else ["Trafik Sigortası", "Kasko", "Sağlık Sigortası"]
        dict_urun = dict(zip(df_urun['Urun_Adi'], df_urun['Komisyon_Orani'])) if not df_urun.empty else {}
        
        df_acente = get_data("Ayarlar_Acenteler")
        acente_listesi = df_acente["Acente_Adi"].tolist() if not df_acente.empty else ["AKTÜRK SİGORTA (MERKEZ)"]
        if "AKTÜRK SİGORTA (MERKEZ)" not in acente_listesi: acente_listesi.insert(0, "AKTÜRK SİGORTA (MERKEZ)")
        dict_acente = dict(zip(df_acente['Acente_Adi'], df_acente['Tali_Oran'])) if not df_acente.empty else {}
        acente_listesi.append("➕ YENİ TALİ ACENTE EKLE")

        st.info("💡 **Yapay Zeka Destekli Okuma:** Müşterinin PDF poliçesini aşağıya sürükleyin, formdaki boşluklar otomatik dolsun.")
        file = st.file_uploader("PDF Poliçe Seçin", type="pdf")
        p_data = {"tanzim":"","baslangic":"","bitis":"","musteri":"","plaka":"","tc_vkn":"","sirket":"","urun":"","p_no":"","net_prim":0.0,"brut_prim":0.0}
        f_bytes = None
        
        if file:
            f_bytes = file.getvalue()
            with st.spinner("📄 PDF Okunuyor..."):
                with pdfplumber.open(io.BytesIO(f_bytes)) as pdf:
                    if txt := pdf.pages[0].extract_text(): 
                        p_data.update(klasik_analiz(txt))
                        st.success("PDF Tarandı! Lütfen boşlukları kontrol edip kaydedin.")

        st.markdown("### 📝 İşlem Türü ve Bağlantı Motoru")
        islem_turu = st.radio("Bu kaydın amacı nedir?", ["Yeni Poliçe / Yenileme", "İptal / Satış", "Zeyl / Teminat Düşürme"], horizontal=True)
        
        ana_pol_data = {}
        if islem_turu != "Yeni Poliçe / Yenileme":
            df_pol_mevcut = get_data("Policeler")
            if not df_pol_mevcut.empty:
                df_pol_mevcut["Ozet"] = df_pol_mevcut["Plaka"] + " | " + df_pol_mevcut["Müşteri Adı Soyadı"] + " | " + df_pol_mevcut["Sigorta Türü"] + " | No: " + df_pol_mevcut["Poliçe No"]
                liste = ["Lütfen Bağlanacak Ana Poliçeyi Seçin..."] + [ozet for ozet in df_pol_mevcut["Ozet"].dropna().unique().tolist() if str(ozet).strip() != ""]
                secim = st.selectbox("🔗 İptal/Zeyil Edilecek Ana Poliçe:", liste)
                if secim != "Lütfen Bağlanacak Ana Poliçeyi Seçin...": ana_pol_data = df_pol_mevcut[df_pol_mevcut["Ozet"] == secim].iloc[0].to_dict()
        
        st.divider()

        with st.form("police_formu", clear_on_submit=True):
            st.subheader("1. Müşteri ve İşlem Detayları")
            c1, c2, c3 = st.columns(3)
            tan = c1.text_input("Tanzim Tarihi", p_data["tanzim"])
            bas = c2.text_input("Başlangıç", p_data["baslangic"])
            bit = c3.text_input("Bitiş", p_data["bitis"])
            
            c4, c5, c6 = st.columns(3)
            def_mus = ana_pol_data.get("Müşteri Adı Soyadı", p_data["musteri"]) if ana_pol_data else p_data["musteri"]
            def_tc = ana_pol_data.get("TC / VKN", p_data["tc_vkn"]) if ana_pol_data else p_data["tc_vkn"]
            mus_girdi = c4.text_input("Müşteri Ad Soyad", def_mus)
            tc = c5.text_input("TC / VKN", def_tc)
            ilet = c6.text_input("Telefon / E-mail", ana_pol_data.get("Telefon / E-mail", ""))
            
            c7, c8, c9 = st.columns(3)
            def_sir = str(ana_pol_data.get("Sigorta Şirketi", p_data["sirket"])).replace(" (İPTAL-SATIŞ)", "").replace(" (İPTAL-ZEYL)", "") if ana_pol_data else p_data["sirket"]
            def_plk = ana_pol_data.get("Plaka", p_data["plaka"]) if ana_pol_data else p_data["plaka"]
            sir_girdi = c7.text_input("Sigorta Şirketi (Doğa, Hepiyi, Allianz vb.)", def_sir)
            pno = c8.text_input("Poliçe No (Bu zeylin/işlemin numarası)", p_data["p_no"])
            plk = c9.text_input("Plaka", def_plk)
            
            st.subheader("2. Finans, Komisyon ve Ödeme")
            c10, c11, c12, c12_kom = st.columns(4)
            def_urun = ana_pol_data.get("Sigorta Türü", p_data["urun"]) if ana_pol_data else p_data["urun"]
            urun_index = urun_listesi.index(def_urun) if def_urun in urun_listesi else 0
            urn = c10.selectbox("Ürün Türü", urun_listesi, index=urun_index)
            net_val = "" if p_data["net_prim"] == 0.0 else str(p_data["net_prim"])
            brut_val = "" if p_data["brut_prim"] == 0.0 else str(p_data["brut_prim"])
            net_girdi = c11.text_input("Net Prim", value=net_val)
            brut_girdi = c12.text_input("Brüt Prim", value=brut_val)
            kom_girdi = c12_kom.text_input("Net Komisyon (Opsiyonel)", value="", help="Boş bırakırsanız ayarlardaki oranı kullanır.")
            
            c13, c14, c15 = st.columns(3)
            acn = c13.selectbox("İşlemi Yapan / Kesilen Ekran", acente_listesi, help="Direkt ekranınızdan (Hepiyi/Doğa) kestiyeniz 'Aktürk Sigorta (Merkez)' seçin.")
            yeni_acente_adi = ""; yeni_acente_orani = 0.0
            if acn == "➕ YENİ TALİ ACENTE EKLE":
                yeni_acente_adi = c13.text_input("Yeni Acente Adını Yazın:")
                yeni_acente_orani_girdi = c13.text_input("Bize Verilen Tali Komisyon Oranı (Örn: 70)", value="")
                yeni_acente_orani = float(sayiya_cevir(yeni_acente_orani_girdi))
                
            odm = c14.selectbox("Müşteri Tahsilat Kanalı", ["Acenteye Borçlanıldı (Cari)", "Nakit Alındı", "Müşteri Kredi Kartı", "Havale"])
            taksit = c15.number_input("Taksit Sayısı", min_value=1, value=1)
            
            c16, c17 = st.columns(2)
            alinan_odeme_str = c16.text_input("Şu An Müşteriden Alınan Ödeme / Peşinat (TL)", value="0", help="Müşteriden peşinat aldıysanız buraya yazın. Müşteri carisinden otomatik düşer.")
            adr = c17.text_area("Adres (Opsiyonel)", ana_pol_data.get("Adres", ""))
            
            st.markdown("<br>", unsafe_allow_html=True)
            if st.form_submit_button("✅ POLİÇEYİ, CARİYİ VE PDF'İ KAYDET", type="primary"):
                with st.spinner("Sisteme İşleniyor..."):
                    def _kaydet_operasyonu():
                        doc = client.open_by_key(SHEET_ID)
                        mus = temiz_isim(mus_girdi)
                        sir = temiz_isim(sir_girdi)
                        aktif_acente = temiz_isim(acn) if acn != "➕ YENİ TALİ ACENTE EKLE" else "➕ YENİ TALİ ACENTE EKLE"
                        plk_temiz = str(plk).replace(" ", "").upper()
                        yeni_fis_no = fis_no_uret("POL")
                        
                        net = float(sayiya_cevir(net_girdi)); brut = float(sayiya_cevir(brut_girdi))
                        kom_manuel = float(sayiya_cevir(kom_girdi)); alinan_odeme = float(sayiya_cevir(alinan_odeme_str))
                        
                        islem_notu = ""; baglanti_notu = f" (Asıl Poliçe No: {ana_pol_data.get('Poliçe No', '')})" if ana_pol_data else ""
                        
                        if islem_turu != "Yeni Poliçe / Yenileme":
                            net = -abs(net); brut = -abs(brut); kom_manuel = -abs(kom_manuel) if kom_manuel > 0 else 0.0
                            if islem_turu == "İptal / Satış": sir_iptal = f"{sir} (İPTAL-SATIŞ)"; islem_notu = f"SATIŞ/TAM İPTAL{baglanti_notu}"
                            else: sir_iptal = f"{sir} (İPTAL-ZEYL)"; islem_notu = f"KISMİ İPTAL/ZEYL{baglanti_notu}"
                        else: sir_iptal = sir
                        
                        link = drive_pdf_yukle(f_bytes, f"{mus}_{plk_temiz}_{sir_iptal}.pdf") if f_bytes else "Yok"
                        u_oran = float(sayiya_cevir(dict_urun.get(urn, 0.0)))
                        if u_oran > 1: u_oran /= 100 
                        
                        sirket_komisyonu = kom_manuel if str(kom_girdi).strip() != "" and kom_manuel != 0.0 else net * u_oran
                        
                        if aktif_acente == "➕ YENİ TALİ ACENTE EKLE" and yeni_acente_adi != "":
                            aktif_acente = temiz_isim(yeni_acente_adi)
                            doc.worksheet("Ayarlar_Acenteler").append_row([aktif_acente, yeni_acente_orani], value_input_option='USER_ENTERED')
                            t_oran = yeni_acente_orani
                        else: t_oran = float(sayiya_cevir(dict_acente.get(aktif_acente, 0.0)))
                        if t_oran > 1: t_oran /= 100 
                        
                        akturk_kazanci = float(sirket_komisyonu * t_oran)
                        islem_tarihi = tarih_formatla(tan)
                        final_pno = pno + baglanti_notu if baglanti_notu else pno
                        
                        ws_pol = doc.worksheet("Policeler")
                        headers = ws_pol.row_values(1)
                        if "Şirket Komisyonu" not in headers: ws_pol.update_cell(1, len(headers)+1, "Şirket Komisyonu"); headers.append("Şirket Komisyonu")
                        
                        row_dict = {"Tanzim Tarihi": islem_tarihi, "Başlangıç Tarihi": bas, "Bitiş Tarihi": bit, "Müşteri Adı Soyadı": mus, "TC / VKN": tc, "Sigorta Şirketi": sir_iptal, "Sigorta Türü": urn, "Poliçe No": final_pno, "Plaka": plk_temiz, "Net Prim": net, "Brüt Prim": brut, "Şirket Komisyonu": sirket_komisyonu, "Acente": aktif_acente, "Adres": adr, "Telefon / E-mail": ilet, "PDF Linki": link}
                        ws_pol.append_row([row_dict.get(h, "") for h in headers], value_input_option='USER_ENTERED')
                        
                        aciklama = f"{sir_iptal.replace(' (İPTAL-SATIŞ)', '').replace(' (İPTAL-ZEYL)', '')} - {urn} - Plaka: {plk_temiz}"
                        yeni_satirlar = []
                        
                        ws_cari = doc.worksheet("Cari_Islemler")
                        c_headers = ws_cari.row_values(1)
                        if "Fiş No" not in c_headers: ws_cari.update_cell(1, len(c_headers)+1, "Fiş No"); c_headers.append("Fiş No")
                            
                        def cari_satir_hazirla(tarih, tur, kurum, detay, borc, alacak, odeme, taksit, fis):
                            satir = [""] * len(c_headers)
                            mapping = {"Tarih": tarih, "Islem_Turu": tur, "Kisi_Kurum": kurum, "Islem_Detayi": detay, "Borc": borc, "Alacak": alacak, "Odeme_Tipi": odeme, "Taksit": taksit, "Fiş No": fis}
                            for key, val in mapping.items():
                                if key in c_headers: satir[c_headers.index(key)] = val
                            return satir

                        if islem_notu: yeni_satirlar.append(cari_satir_hazirla(islem_tarihi, "Müşteri Carisi", mus, f"{islem_notu} İADESİ - {aciklama}", brut, 0.0, odm, taksit, yeni_fis_no))
                        else: yeni_satirlar.append(cari_satir_hazirla(islem_tarihi, "Müşteri Carisi", mus, aciklama, brut, 0.0, odm, taksit, yeni_fis_no))
                        if alinan_odeme > 0: yeni_satirlar.append(cari_satir_hazirla(islem_tarihi, "Müşteri Carisi", mus, f"Anında Tahsilat - {aciklama}", 0.0, alinan_odeme, odm, 1, yeni_fis_no))
                        
                        if aktif_acente == "AKTÜRK SİGORTA (MERKEZ)":
                            sirket_adi_temiz = sir_iptal.replace(' (İPTAL-SATIŞ)', '').replace(' (İPTAL-ZEYL)', '').strip()
                            if islem_notu: yeni_satirlar.append(cari_satir_hazirla(islem_tarihi, "Sigorta Şirketi Carisi", sirket_adi_temiz, f"Şirket Komisyonu İptal/Kesintisi - {aciklama}{baglanti_notu}", sirket_komisyonu, 0.0, "Belge Bekleniyor", 1, yeni_fis_no))
                            else: yeni_satirlar.append(cari_satir_hazirla(islem_tarihi, "Sigorta Şirketi Carisi", sirket_adi_temiz, f"Şirket Komisyonu Hakediş - {aciklama}", sirket_komisyonu, 0.0, "Belge Bekleniyor", 1, yeni_fis_no))
                        else:
                            if islem_notu: yeni_satirlar.append(cari_satir_hazirla(islem_tarihi, "Tali Acente Carisi", aktif_acente, f"Acente Payı İptal/Kesintisi - {aciklama}{baglanti_notu}", akturk_kazanci, 0.0, "Fatura Bekleniyor", 1, yeni_fis_no))
                            else: yeni_satirlar.append(cari_satir_hazirla(islem_tarihi, "Tali Acente Carisi", aktif_acente, f"Acente Payı Hakediş - {aciklama}", akturk_kazanci, 0.0, "Fatura Bekleniyor", 1, yeni_fis_no))
                        
                        ws_cari.append_rows(yeni_satirlar, value_input_option='USER_ENTERED')
                        doc.worksheet("Musteriler").append_row([mus, tc, ilet, brut], value_input_option='USER_ENTERED')
                        return True
                    
                    if api_kalkani(_kaydet_operasyonu) is not None: st.success("🎉 Harika! Poliçe başarıyla kaydedildi ve tüm cari hesaplara hatasız işlendi."); st.cache_data.clear()

    elif menu == "💰 Cari & Mutabakat":
        st.header("💰 Finans ve Mutabakat Yönetimi")
        
        t_merkez, t1, t2, t3, t4 = st.tabs(["🏦 AKTÜRK MERKEZ KASA", "🏢 Tali Acente Hesapları", "🏛️ Sigorta Şirketi Hesapları", "👤 Müşteri Hesapları", "📊 Toplu Bilançolar"])
        
        df_pol = get_data("Policeler"); df_cari = get_data("Cari_Islemler")
        df_urunler = get_data("Ayarlar_Urunler"); df_acenteler = get_data("Ayarlar_Acenteler"); df_sirketler = get_data("Ayarlar_Sirketler")

        # ----------------------------------------
        # YENİ SEKME: AKTÜRK SİGORTA KASA KONSOLİDE
        # ----------------------------------------
        with t_merkez:
            st.markdown("### 🏦 Aktürk Sigorta Merkez Kasa (Konsolide Komisyon ve Tahsilat)")
            st.info("💡 **Burada tüm sigorta şirketlerinden (Doğa, Hepiyi vb.) ve tüm tali acentelerden (Edy, Sigortam Güvende vb.) alacağınız toplam komisyonları ve hesabınıza yatan (tahsil edilen) paraları tek bir havuzda görürsünüz.**")
            
            c_tarih1, c_tarih2 = st.columns(2)
            ilk_tarih_m = c_tarih1.date_input("Kasa Başlangıç Tarihi", datetime.today().replace(month=1, day=1), key="merkez_bas")
            son_tarih_m = c_tarih2.date_input("Kasa Bitiş Tarihi", datetime.today(), key="merkez_bit")
            st.divider()

            if not df_cari.empty:
                merkez_df = df_cari[df_cari["Islem_Turu"].isin(["Tali Acente Carisi", "Sigorta Şirketi Carisi"])].copy()
                if not merkez_df.empty:
                    merkez_df["Borc"] = merkez_df["Borc"].apply(sayiya_cevir)
                    merkez_df["Alacak"] = merkez_df["Alacak"].apply(sayiya_cevir)
                    merkez_df['Tarih_Obj'] = pd.to_datetime(merkez_df['Tarih'], dayfirst=True, errors='coerce')
                    
                    gecmis_mask = merkez_df['Tarih_Obj'].dt.date < ilk_tarih_m
                    devir_hakedis = merkez_df[gecmis_mask]["Borc"].sum()
                    devir_tahsilat = merkez_df[gecmis_mask]["Alacak"].sum()
                    devir_bakiye = devir_hakedis - devir_tahsilat
                    
                    devir_satiri = pd.DataFrame()
                    if devir_bakiye != 0:
                        devir_satiri = pd.DataFrame([{
                            "Tarih": ilk_tarih_m.strftime("%d.%m.%Y"), "Hesap_Turu": "-", "Kurum": "🔄 ÖNCEKİ DÖNEMDEN DEVİR", "İşlem Detayı": "-", "Fiş No": "-",
                            "Hakediş (Bize Borç)": devir_bakiye if devir_bakiye > 0 else 0.0, "Tahsilat (Bize Ödenen)": abs(devir_bakiye) if devir_bakiye < 0 else 0.0
                        }])

                    mask_cari = (merkez_df['Tarih_Obj'].dt.date >= ilk_tarih_m) & (merkez_df['Tarih_Obj'].dt.date <= son_tarih_m)
                    donem_df = merkez_df[mask_cari].copy()
                    
                    donem_df.rename(columns={"Islem_Turu": "Hesap_Turu", "Kisi_Kurum": "Kurum", "Islem_Detayi": "İşlem Detayı", "Borc": "Hakediş (Bize Borç)", "Alacak": "Tahsilat (Bize Ödenen)"}, inplace=True)
                    gosterilecek_merkez_df = donem_df[["Tarih", "Hesap_Turu", "Kurum", "Fiş No", "İşlem Detayı", "Hakediş (Bize Borç)", "Tahsilat (Bize Ödenen)"]]
                    if not devir_satiri.empty: gosterilecek_merkez_df = pd.concat([devir_satiri, gosterilecek_merkez_df], ignore_index=True)
                        
                    gosterilecek_merkez_df["Bakiye (Kalan)"] = gosterilecek_merkez_df["Hakediş (Bize Borç)"].astype(float).cumsum() - gosterilecek_merkez_df["Tahsilat (Bize Ödenen)"].astype(float).cumsum()
                    
                    total_hakedis = merkez_df["Borc"].sum(); total_tahsilat = merkez_df["Alacak"].sum(); total_kalan = total_hakedis - total_tahsilat
                    
                    m1, m2, m3 = st.columns(3)
                    m1.metric("📊 Toplam Komisyon Hakedişi", para_format(total_hakedis))
                    m2.metric("💰 Toplam Kasaya Giren (Tahsilat)", para_format(total_tahsilat))
                    m3.metric("⏳ Piyasadan Kalan Alacak", para_format(total_kalan))
                    
                    df_ui_merkez = df_gorsel_yap(gosterilecek_merkez_df, ["Hakediş (Bize Borç)", "Tahsilat (Bize Ödenen)", "Bakiye (Kalan)"])
                    st.dataframe(df_ui_merkez, use_container_width=True)
                    excel_indir(gosterilecek_merkez_df, "Kasa ve Konsolide Ekstresini İndir", "Akturk_Merkez_Kasa_Ekstresi")

        # ----------------------------------------
        # SEKME 1: TALİ ACENTE HESAPLARI
        # ----------------------------------------
        with t1:
            acente_adlari = df_acenteler["Acente_Adi"].dropna().unique().tolist() if not df_acenteler.empty else []
            eski_taliler = df_cari[df_cari["Islem_Turu"] == "Tali Acente Carisi"]["Kisi_Kurum"].dropna().unique().tolist() if not df_cari.empty and "Kisi_Kurum" in df_cari.columns else []
            # Merkez hesabı Tali Listesinden çıkartıldı, o artık patron hesabı.
            tali_listesi = sorted([x for x in list(set(acente_adlari + eski_taliler)) if str(x).strip() != "" and x != "AKTÜRK SİGORTA (MERKEZ)"])
            
            secilen_tali = st.selectbox("📌 Mutabakat Yapılacak Tali Acente:", ["Seçiniz..."] + tali_listesi, key="sel_tali")
            
            if secilen_tali != "Seçiniz...":
                c_tarih1, c_tarih2 = st.columns(2)
                ilk_tarih = c_tarih1.date_input("Başlangıç Tarihi", datetime.today().replace(month=1, day=1), key="tali_bas")
                son_tarih = c_tarih2.date_input("Bitiş Tarihi", datetime.today(), key="tali_bit")
                st.divider()

                st.markdown(f"#### 1️⃣ {secilen_tali} - Tali Acente Poliçe Üretimi")
                if not df_pol.empty:
                    kurum_policeleri = df_pol[df_pol["Acente"] == secilen_tali].copy()
                    if not kurum_policeleri.empty:
                        kurum_policeleri['Tarih_Obj'] = pd.to_datetime(kurum_policeleri['Tanzim Tarihi'], dayfirst=True, errors='coerce')
                        mask_pol = (kurum_policeleri['Tarih_Obj'].dt.date >= ilk_tarih) & (kurum_policeleri['Tarih_Obj'].dt.date <= son_tarih)
                        filtrelenmis_pol = kurum_policeleri[mask_pol]
                        
                        if not filtrelenmis_pol.empty:
                            m_col1, m_col2, m_col3 = st.columns(3)
                            m_col1.metric("Toplam İşlem Adedi", f"{len(filtrelenmis_pol)} Poliçe")
                            m_col2.metric("Brüt Üretim (TL)", para_format(filtrelenmis_pol["Brüt Prim"].apply(sayiya_cevir).sum()))
                            m_col3.metric("Net Üretim (TL)", para_format(filtrelenmis_pol["Net Prim"].apply(sayiya_cevir).sum()))
                            
                            goster_pol = df_gorsel_yap(filtrelenmis_pol[["Tanzim Tarihi", "Müşteri Adı Soyadı", "Plaka", "Sigorta Şirketi", "Sigorta Türü", "Net Prim", "Brüt Prim", "Şirket Komisyonu"]], ["Net Prim", "Brüt Prim", "Şirket Komisyonu"])
                            st.dataframe(goster_pol, use_container_width=True)
                            excel_indir(filtrelenmis_pol, "Poliçe Üretimlerini Excel İndir", f"{secilen_tali}_Uretim_Listesi")
                        else: st.warning("Bu tarih aralığında poliçe kesilmemiş.")
                    else: st.warning("Poliçe kaydı bulunamadı.")
                
                st.divider()

                st.markdown(f"#### 2️⃣ {secilen_tali} - Finansal Mutabakat (Tali Acente Carisi)")
                if not df_cari.empty:
                    kurum_carisi = df_cari[(df_cari["Kisi_Kurum"] == secilen_tali) & (df_cari["Islem_Turu"] == "Tali Acente Carisi")].copy()
                    if not kurum_carisi.empty:
                        kurum_carisi = hesap_kodu_ekle(kurum_carisi, "Tali")
                        kurum_carisi["Borc"] = kurum_carisi["Borc"].apply(sayiya_cevir); kurum_carisi["Alacak"] = kurum_carisi["Alacak"].apply(sayiya_cevir)
                        kurum_carisi['Tarih_Obj'] = pd.to_datetime(kurum_carisi['Tarih'], dayfirst=True, errors='coerce')
                        
                        genel_bakiye = kurum_carisi["Borc"].sum() - kurum_carisi["Alacak"].sum()
                        if genel_bakiye > 0: st.error(f"🚨 Tali Acentenin Bize Borcu: {para_format(genel_bakiye)}")
                        elif genel_bakiye < 0: st.success(f"✅ Sizin Tali Acenteye Borcunuz: {para_format(abs(genel_bakiye))}")
                        else: st.info("✅ Hesaplar Mutabık (0,00 TL)")
                        
                        gecmis_mask = kurum_carisi['Tarih_Obj'].dt.date < ilk_tarih
                        devir_bakiye = kurum_carisi[gecmis_mask]["Borc"].sum() - kurum_carisi[gecmis_mask]["Alacak"].sum()
                        
                        devir_satiri = pd.DataFrame()
                        if devir_bakiye != 0:
                            devir_satiri = pd.DataFrame([{"Hesap Kodu": kurum_carisi["Hesap Kodu"].iloc[0] if not kurum_carisi.empty else "", "Tarih": ilk_tarih.strftime("%d.%m.%Y"), "Fiş No": "-", "Islem_Detayi": "🔄 ÖNCEKİ DÖNEMDEN DEVİR", "Borc": devir_bakiye if devir_bakiye > 0 else 0.0, "Alacak": abs(devir_bakiye) if devir_bakiye < 0 else 0.0, "Odeme_Tipi": "-"}])

                        mask_cari = (kurum_carisi['Tarih_Obj'].dt.date >= ilk_tarih) & (kurum_carisi['Tarih_Obj'].dt.date <= son_tarih)
                        gosterilecek_kolonlar = ["Hesap Kodu", "Tarih", "Fiş No", "Islem_Detayi", "Borc", "Alacak", "Odeme_Tipi"]
                        gosterilecek_cari_df = kurum_carisi[mask_cari][[c for c in gosterilecek_kolonlar if c in kurum_carisi.columns]]
                        if not devir_satiri.empty: gosterilecek_cari_df = pd.concat([devir_satiri, gosterilecek_cari_df], ignore_index=True)
                        
                        gosterilecek_cari_df["Bakiye (Kalan)"] = gosterilecek_cari_df["Borc"].astype(float).cumsum() - gosterilecek_cari_df["Alacak"].astype(float).cumsum()
                        
                        df_ui_cari = df_gorsel_yap(gosterilecek_cari_df, ["Borc", "Alacak", "Bakiye (Kalan)"])
                        st.dataframe(df_ui_cari, use_container_width=True)
                        excel_indir(gosterilecek_cari_df, "Tali Ekstreyi Excel İndir", f"{secilen_tali}_Ekstre")

                st.divider()
                with st.form("tali_islem", clear_on_submit=True):
                    st.markdown(f"💳 **{secilen_tali} - Tahsilat ve Fatura İşlemi**")
                    c1, c2, c3, c4 = st.columns(4)
                    o_tarih = c1.date_input("İşlem Tarihi", key="t_tarih_tali").strftime("%d.%m.%Y")
                    islem_yonu = c2.selectbox("İşlem Türü", ["💰 Ödeme Geldi (Bize Yapılan Tahsilat)", "💸 Ödeme Çıktı (Bizden Kuruma Yapılan)", "📄 Fatura / Belge Kesildi"], key="t_yon_tali")
                    o_tutar = float(sayiya_cevir(c3.text_input("Tutar (Örn: 1500,50)", value="", key="t_tutar_tali")))
                    o_detay = c4.text_input("Açıklama", key="t_detay_tali")
                    
                    if st.form_submit_button("💾 Tali İşlemini Kaydet", type="primary"):
                        with st.spinner("İşleniyor..."):
                            def _tali_odeme():
                                ws_cari = client.open_by_key(SHEET_ID).worksheet("Cari_Islemler")
                                c_headers = ws_cari.row_values(1)
                                if "Fiş No" not in c_headers: ws_cari.update_cell(1, len(c_headers)+1, "Fiş No"); c_headers.append("Fiş No")
                                satir = [""] * len(c_headers)
                                mapping = {"Tarih": o_tarih, "Islem_Turu": "Tali Acente Carisi", "Kisi_Kurum": secilen_tali, "Islem_Detayi": f"{islem_yonu} - {o_detay}", "Borc": o_tutar if "Ödeme Çıktı" in islem_yonu else 0.0, "Alacak": o_tutar if "Ödeme Geldi" in islem_yonu else 0.0, "Odeme_Tipi": "Banka/Resmi Evrak", "Taksit": 1, "Fiş No": fis_no_uret("THS")}
                                for key, val in mapping.items():
                                    if key in c_headers: satir[c_headers.index(key)] = val
                                ws_cari.append_row(satir, value_input_option='USER_ENTERED')
                                return True
                            if api_kalkani(_tali_odeme): st.cache_data.clear(); st.rerun()

        # ----------------------------------------
        # SEKME 2: SİGORTA ŞİRKETİ HESAPLARI
        # ----------------------------------------
        with t2:
            sirket_adlari = df_sirketler["Sirket_Adi"].dropna().unique().tolist() if not df_sirketler.empty and "Sirket_Adi" in df_sirketler.columns else []
            eski_sirketler = df_cari[df_cari["Islem_Turu"] == "Sigorta Şirketi Carisi"]["Kisi_Kurum"].dropna().unique().tolist() if not df_cari.empty and "Kisi_Kurum" in df_cari.columns else []
            sirket_listesi = sorted([x for x in list(set(sirket_adlari + eski_sirketler)) if str(x).strip() != ""])
            
            secilen_sirket = st.selectbox("📌 Mutabakat Yapılacak Sigorta Şirketi:", ["Seçiniz..."] + sirket_listesi, key="sel_sirket")
            
            if secilen_sirket != "Seçiniz...":
                c_tarih1, c_tarih2 = st.columns(2)
                ilk_tarih = c_tarih1.date_input("Başlangıç Tarihi", datetime.today().replace(month=1, day=1), key="sirket_bas")
                son_tarih = c_tarih2.date_input("Bitiş Tarihi", datetime.today(), key="sirket_bit")
                st.divider()

                st.markdown(f"#### 1️⃣ {secilen_sirket} - Direkt Ekran Poliçe Üretimi")
                if not df_pol.empty:
                    kurum_policeleri = df_pol[df_pol["Sigorta Şirketi"].str.replace(r' \(İPTAL-.*?\)', '', regex=True).str.strip() == secilen_sirket].copy()
                    if not kurum_policeleri.empty:
                        kurum_policeleri['Tarih_Obj'] = pd.to_datetime(kurum_policeleri['Tanzim Tarihi'], dayfirst=True, errors='coerce')
                        mask_pol = (kurum_policeleri['Tarih_Obj'].dt.date >= ilk_tarih) & (kurum_policeleri['Tarih_Obj'].dt.date <= son_tarih)
                        filtrelenmis_pol = kurum_policeleri[mask_pol]
                        
                        if not filtrelenmis_pol.empty:
                            m_col1, m_col2, m_col3 = st.columns(3)
                            m_col1.metric("Toplam İşlem Adedi", f"{len(filtrelenmis_pol)} Poliçe")
                            m_col2.metric("Brüt Üretim (TL)", para_format(filtrelenmis_pol["Brüt Prim"].apply(sayiya_cevir).sum()))
                            m_col3.metric("Net Üretim (TL)", para_format(filtrelenmis_pol["Net Prim"].apply(sayiya_cevir).sum()))
                            
                            goster_pol = df_gorsel_yap(filtrelenmis_pol[["Tanzim Tarihi", "Müşteri Adı Soyadı", "Plaka", "Sigorta Şirketi", "Sigorta Türü", "Net Prim", "Brüt Prim", "Şirket Komisyonu"]], ["Net Prim", "Brüt Prim", "Şirket Komisyonu"])
                            st.dataframe(goster_pol, use_container_width=True)
                            excel_indir(filtrelenmis_pol, "Poliçe Üretimlerini Excel İndir", f"{secilen_sirket}_Uretim_Listesi")
                        else: st.warning("Bu tarih aralığında poliçe kesilmemiş.")
                    else: st.warning("Poliçe kaydı bulunamadı.")
                
                st.divider()

                st.markdown(f"#### 2️⃣ {secilen_sirket} - Finansal Mutabakat (Şirket Carisi)")
                if not df_cari.empty:
                    kurum_carisi = df_cari[(df_cari["Kisi_Kurum"] == secilen_sirket) & (df_cari["Islem_Turu"] == "Sigorta Şirketi Carisi")].copy()
                    if not kurum_carisi.empty:
                        kurum_carisi = hesap_kodu_ekle(kurum_carisi, "Şirket")
                        kurum_carisi["Borc"] = kurum_carisi["Borc"].apply(sayiya_cevir); kurum_carisi["Alacak"] = kurum_carisi["Alacak"].apply(sayiya_cevir)
                        kurum_carisi['Tarih_Obj'] = pd.to_datetime(kurum_carisi['Tarih'], dayfirst=True, errors='coerce')
                        
                        genel_bakiye = kurum_carisi["Borc"].sum() - kurum_carisi["Alacak"].sum()
                        if genel_bakiye > 0: st.error(f"🚨 Sigorta Şirketinin Bize Borcu (Alacağımız): {para_format(genel_bakiye)}")
                        elif genel_bakiye < 0: st.success(f"✅ Sizin Sigorta Şirketine Borcunuz: {para_format(abs(genel_bakiye))}")
                        else: st.info("✅ Hesaplar Mutabık (0,00 TL)")
                        
                        gecmis_mask = kurum_carisi['Tarih_Obj'].dt.date < ilk_tarih
                        devir_bakiye = kurum_carisi[gecmis_mask]["Borc"].sum() - kurum_carisi[gecmis_mask]["Alacak"].sum()
                        
                        devir_satiri = pd.DataFrame()
                        if devir_bakiye != 0:
                            devir_satiri = pd.DataFrame([{"Hesap Kodu": kurum_carisi["Hesap Kodu"].iloc[0] if not kurum_carisi.empty else "", "Tarih": ilk_tarih.strftime("%d.%m.%Y"), "Fiş No": "-", "Islem_Detayi": "🔄 ÖNCEKİ DÖNEMDEN DEVİR", "Borc": devir_bakiye if devir_bakiye > 0 else 0.0, "Alacak": abs(devir_bakiye) if devir_bakiye < 0 else 0.0, "Odeme_Tipi": "-"}])

                        mask_cari = (kurum_carisi['Tarih_Obj'].dt.date >= ilk_tarih) & (kurum_carisi['Tarih_Obj'].dt.date <= son_tarih)
                        gosterilecek_kolonlar = ["Hesap Kodu", "Tarih", "Fiş No", "Islem_Detayi", "Borc", "Alacak", "Odeme_Tipi"]
                        gosterilecek_cari_df = kurum_carisi[mask_cari][[c for c in gosterilecek_kolonlar if c in kurum_carisi.columns]]
                        if not devir_satiri.empty: gosterilecek_cari_df = pd.concat([devir_satiri, gosterilecek_cari_df], ignore_index=True)
                        
                        gosterilecek_cari_df["Bakiye (Kalan)"] = gosterilecek_cari_df["Borc"].astype(float).cumsum() - gosterilecek_cari_df["Alacak"].astype(float).cumsum()

                        df_ui_cari = df_gorsel_yap(gosterilecek_cari_df, ["Borc", "Alacak", "Bakiye (Kalan)"])
                        st.dataframe(df_ui_cari, use_container_width=True)
                        excel_indir(gosterilecek_cari_df, "Şirket Ekstresini Excel İndir", f"{secilen_sirket}_Ekstre")

                st.divider()
                with st.form("sirket_islem", clear_on_submit=True):
                    st.markdown(f"💳 **{secilen_sirket} - Tahsilat ve Fatura İşlemi**")
                    c1, c2, c3, c4 = st.columns(4)
                    o_tarih = c1.date_input("İşlem Tarihi", key="s_tarih_sirket").strftime("%d.%m.%Y")
                    islem_yonu = c2.selectbox("İşlem Türü", ["💰 Ödeme Geldi (Bize Yapılan Tahsilat)", "💸 Ödeme Çıktı (Bizden Kuruma Yapılan)", "📄 Fatura / Belge Kesildi"], key="s_yon_sirket")
                    o_tutar = float(sayiya_cevir(c3.text_input("Tutar (Örn: 1500,50)", value="", key="s_tutar_sirket")))
                    o_detay = c4.text_input("Açıklama", key="s_detay_sirket")
                    
                    if st.form_submit_button("💾 Şirket İşlemini Kaydet", type="primary"):
                        with st.spinner("İşleniyor..."):
                            def _sirket_odeme():
                                ws_cari = client.open_by_key(SHEET_ID).worksheet("Cari_Islemler")
                                c_headers = ws_cari.row_values(1)
                                if "Fiş No" not in c_headers: ws_cari.update_cell(1, len(c_headers)+1, "Fiş No"); c_headers.append("Fiş No")
                                satir = [""] * len(c_headers)
                                mapping = {"Tarih": o_tarih, "Islem_Turu": "Sigorta Şirketi Carisi", "Kisi_Kurum": secilen_sirket, "Islem_Detayi": f"{islem_yonu} - {o_detay}", "Borc": o_tutar if "Ödeme Çıktı" in islem_yonu else 0.0, "Alacak": o_tutar if "Ödeme Geldi" in islem_yonu else 0.0, "Odeme_Tipi": "Banka/Resmi Evrak", "Taksit": 1, "Fiş No": fis_no_uret("THS")}
                                for key, val in mapping.items():
                                    if key in c_headers: satir[c_headers.index(key)] = val
                                ws_cari.append_row(satir, value_input_option='USER_ENTERED')
                                return True
                            if api_kalkani(_sirket_odeme): st.cache_data.clear(); st.rerun()

        # ----------------------------------------
        # SEKME 3: MÜŞTERİ HESAPLARI
        # ----------------------------------------
        with t3:
            if not df_cari.empty and "Kisi_Kurum" in df_cari.columns:
                musteriler = [m for m in df_cari[df_cari["Islem_Turu"] == "Müşteri Carisi"]["Kisi_Kurum"].dropna().unique().tolist() if str(m).strip() != ""]
                secilen_musteri = st.selectbox("📌 Müşteri Seçin:", ["Seçiniz..."] + sorted(musteriler))
                if secilen_musteri != "Seçiniz...":
                    m_cari = df_cari[(df_cari["Kisi_Kurum"] == secilen_musteri) & (df_cari["Islem_Turu"] == "Müşteri Carisi")].copy()
                    m_cari = hesap_kodu_ekle(m_cari, "Müşteri")
                    
                    c_mtarih1, c_mtarih2 = st.columns(2)
                    m_ilk = c_mtarih1.date_input("Müşteri Başlangıç", datetime.today().replace(month=1, day=1))
                    m_son = c_mtarih2.date_input("Müşteri Bitiş", datetime.today())
                    
                    m_cari["Borc"] = m_cari["Borc"].apply(sayiya_cevir); m_cari["Alacak"] = m_cari["Alacak"].apply(sayiya_cevir)
                    m_cari['Tarih_Obj'] = pd.to_datetime(m_cari['Tarih'], dayfirst=True, errors='coerce')
                    
                    bakiye = m_cari["Borc"].sum() - m_cari["Alacak"].sum()
                    if bakiye > 0: st.error(f"🚨 Müşterinin Toplam Borcu: {para_format(bakiye)}")
                    elif bakiye < 0: st.success(f"✅ Fazla Ödeme (Alacaklı): {para_format(abs(bakiye))}")
                    else: st.info("✅ Borç Yok (0,00 TL)")
                    
                    gecmis_m_mask = m_cari['Tarih_Obj'].dt.date < m_ilk
                    devir_m_bakiye = m_cari[gecmis_m_mask]["Borc"].sum() - m_cari[gecmis_m_mask]["Alacak"].sum()
                    
                    devir_m_satiri = pd.DataFrame()
                    if devir_m_bakiye != 0:
                        devir_m_satiri = pd.DataFrame([{"Hesap Kodu": m_cari["Hesap Kodu"].iloc[0] if not m_cari.empty else "", "Tarih": m_ilk.strftime("%d.%m.%Y"), "Fiş No": "-", "Islem_Detayi": "🔄 ÖNCEKİ DÖNEMDEN DEVİR", "Borc": devir_m_bakiye if devir_m_bakiye > 0 else 0.0, "Alacak": abs(devir_m_bakiye) if devir_m_bakiye < 0 else 0.0, "Odeme_Tipi": "-"}])

                    mask_m = (m_cari['Tarih_Obj'].dt.date >= m_ilk) & (m_cari['Tarih_Obj'].dt.date <= m_son)
                    gosterilecek_kolonlar = ["Hesap Kodu", "Tarih", "Fiş No", "Islem_Detayi", "Borc", "Alacak", "Odeme_Tipi"]
                    gosterilecek_m_df = m_cari[mask_m][[c for c in gosterilecek_kolonlar if c in m_cari.columns]]
                    if not devir_m_satiri.empty: gosterilecek_m_df = pd.concat([devir_m_satiri, gosterilecek_m_df], ignore_index=True)
                    
                    gosterilecek_m_df["Bakiye (Kalan)"] = gosterilecek_m_df["Borc"].astype(float).cumsum() - gosterilecek_m_df["Alacak"].astype(float).cumsum()

                    df_ui_mus = df_gorsel_yap(gosterilecek_m_df, ["Borc", "Alacak", "Bakiye (Kalan)"])
                    st.dataframe(df_ui_mus, use_container_width=True)
                    excel_indir(gosterilecek_m_df, "Müşteri Ekstresini İndir", f"{secilen_musteri}_Ekstre")

                    st.divider()
                    with st.form("musteri_odeme", clear_on_submit=True):
                        st.markdown(f"💳 **{secilen_musteri} - Tahsilat / İade Girişi**")
                        c1, c2, c3, c4 = st.columns(4)
                        m_tarih = c1.date_input("Tarih").strftime("%d.%m.%Y")
                        m_yon = c2.selectbox("Tür", ["Müşteriden Para Geldi (Tahsilat)", "Müşteriye Para İade Edildi"])
                        m_tutar = float(sayiya_cevir(c3.text_input("Tutar (Örn: 1500,50)", value="")))
                        m_detay = c4.text_input("Açıklama")
                        if st.form_submit_button("💾 İşlemi Kaydet", type="primary"):
                            with st.spinner("İşleniyor, lütfen bekleyiniz..."):
                                def _mus_odeme():
                                    ws_cari = client.open_by_key(SHEET_ID).worksheet("Cari_Islemler")
                                    c_headers = ws_cari.row_values(1)
                                    if "Fiş No" not in c_headers: ws_cari.update_cell(1, len(c_headers)+1, "Fiş No"); c_headers.append("Fiş No")
                                    satir = [""] * len(c_headers)
                                    mapping = {"Tarih": m_tarih, "Islem_Turu": "Müşteri Carisi", "Kisi_Kurum": secilen_musteri, "Islem_Detayi": m_detay, "Borc": m_tutar if m_yon == "Müşteriye Para İade Edildi" else 0.0, "Alacak": m_tutar if m_yon == "Müşteriden Para Geldi (Tahsilat)" else 0.0, "Odeme_Tipi": "Nakit/Havale", "Taksit": 1, "Fiş No": fis_no_uret("THS")}
                                    for key, val in mapping.items():
                                        if key in c_headers: satir[c_headers.index(key)] = val
                                    ws_cari.append_row(satir, value_input_option='USER_ENTERED')
                                    return True
                                if api_kalkani(_mus_odeme) is not None: st.cache_data.clear(); st.rerun()

        # ----------------------------------------
        # SEKME 4: TOPLU BİLANÇOLAR VE GRAFİKLER
        # ----------------------------------------
        with t4:
            sub1, sub2, sub3, sub4 = st.tabs(["📋 Müşteri Bilançoları", "📋 Tali Acente Bilançoları", "📋 Sigorta Şirketi Bilançoları", "📈 Gelir ve Üretim Analizi"])
            if not df_cari.empty:
                df_ozet = df_cari[df_cari["Kisi_Kurum"].str.strip() != ""].copy()
                df_ozet["Borc"] = df_ozet["Borc"].apply(sayiya_cevir); df_ozet["Alacak"] = df_ozet["Alacak"].apply(sayiya_cevir)
                grup = df_ozet.groupby(["Islem_Turu", "Kisi_Kurum"])[["Borc", "Alacak"]].sum().reset_index()
                grup["Bakiye"] = grup["Borc"] - grup["Alacak"]
                grup.rename(columns={"Islem_Turu": "Hesap Türü", "Kisi_Kurum": "Kişi / Kurum / Acente", "Borc": "Toplam Borç", "Alacak": "Toplam Alacak", "Bakiye": "Dip Tutar (Bakiye)"}, inplace=True)
                
                with sub1:
                    df_musteri_bilanco = grup[grup["Hesap Türü"] == "Müşteri Carisi"].copy()
                    df_musteri_bilanco = hesap_kodu_ekle(df_musteri_bilanco, "Müşteri").sort_values(by="Dip Tutar (Bakiye)", ascending=False)
                    st.dataframe(df_gorsel_yap(df_musteri_bilanco, ["Toplam Borç", "Toplam Alacak", "Dip Tutar (Bakiye)"]), use_container_width=True)
                    excel_indir(df_musteri_bilanco, "Müşteri Bilançosunu Excel İndir", "Musteri_Bilanco_Raporu")

                with sub2:
                    df_tali_bilanco = grup[grup["Hesap Türü"] == "Tali Acente Carisi"].copy()
                    df_tali_bilanco = hesap_kodu_ekle(df_tali_bilanco, "Tali").sort_values(by="Dip Tutar (Bakiye)", ascending=False)
                    st.dataframe(df_gorsel_yap(df_tali_bilanco, ["Toplam Borç", "Toplam Alacak", "Dip Tutar (Bakiye)"]), use_container_width=True)
                    excel_indir(df_tali_bilanco, "Tali Bilançosunu Excel İndir", "Tali_Bilanco_Raporu")
                    
                with sub3:
                    df_sirket_bilanco = grup[grup["Hesap Türü"] == "Sigorta Şirketi Carisi"].copy()
                    df_sirket_bilanco = hesap_kodu_ekle(df_sirket_bilanco, "Şirket").sort_values(by="Dip Tutar (Bakiye)", ascending=False)
                    st.dataframe(df_gorsel_yap(df_sirket_bilanco, ["Toplam Borç", "Toplam Alacak", "Dip Tutar (Bakiye)"]), use_container_width=True)
                    excel_indir(df_sirket_bilanco, "Şirket Bilançosunu Excel İndir", "Sirket_Bilanco_Raporu")
                    
            with sub4:
                st.markdown("### 📈 Dönemsel Gelir ve Üretim Grafikleri")
                c_f1, c_f2 = st.columns(2)
                g_basla = c_f1.date_input("Analiz Başlangıç Tarihi", datetime.today().replace(day=1, month=1))
                g_bitis = c_f2.date_input("Analiz Bitiş Tarihi", datetime.today())
                if not df_pol.empty:
                    df_g = df_pol.copy()
                    df_g['Tarih_Obj'] = pd.to_datetime(df_g['Tanzim Tarihi'], dayfirst=True, errors='coerce')
                    df_g = df_g[(df_g['Tarih_Obj'].dt.date >= g_basla) & (df_g['Tarih_Obj'].dt.date <= g_bitis)]
                    if not df_g.empty:
                        tum_acenteler = ["Tüm Acenteler (Genel)"] + df_g["Acente"].dropna().unique().tolist()
                        tum_urunler = ["Tüm Ürünler (Genel)"] + df_g["Sigorta Türü"].dropna().unique().tolist()
                        c_f3, c_f4 = st.columns(2)
                        sec_acente_g = c_f3.selectbox("Acente Filtresi", tum_acenteler)
                        sec_urun_g = c_f4.selectbox("Ürün Türü Filtresi (Trafik, Kasko vb.)", tum_urunler)
                        
                        if sec_acente_g != "Tüm Acenteler (Genel)": df_g = df_g[df_g["Acente"] == sec_acente_g]
                        if sec_urun_g != "Tüm Ürünler (Genel)": df_g = df_g[df_g["Sigorta Türü"] == sec_urun_g]
                        
                        if not df_g.empty:
                            df_g["Net Prim"] = df_g["Net Prim"].apply(sayiya_cevir); df_g["Brüt Prim"] = df_g["Brüt Prim"].apply(sayiya_cevir)
                            def calc_kazanc(row):
                                kom = sayiya_cevir(row.get("Şirket Komisyonu", 0))
                                if kom == 0:
                                    u_or = float(sayiya_cevir(urun_oranlari.get(row.get("Sigorta Türü", ""), 0.0)))
                                    kom = row["Net Prim"] * (u_or / 100 if u_or > 1 else u_or)
                                acn = row.get("Acente", "AKTÜRK SİGORTA (MERKEZ)")
                                if acn == "AKTÜRK SİGORTA (MERKEZ)": return kom
                                else:
                                    t_or = float(sayiya_cevir(acente_oranlari.get(acn, 0.0)))
                                    return kom * (t_or / 100 if t_or > 1 else t_or)
                                    
                            df_g["Aktürk Sigorta Kazancı"] = df_g.apply(calc_kazanc, axis=1)
                            df_g["Ay-Yıl"] = df_g["Tarih_Obj"].dt.strftime('%Y-%m')
                            
                            m1, m2, m3 = st.columns(3)
                            m1.metric("📊 Toplam Net Üretim", para_format(df_g["Net Prim"].sum()))
                            m2.metric("💰 Toplam Aktürk Kazancı", para_format(df_g["Aktürk Sigorta Kazancı"].sum()))
                            m3.metric("📝 Toplam İşlem Adedi", len(df_g))
                            st.divider()
                            st.markdown("#### 📅 Aylık Üretim ve Kazanç Trendi")
                            grup_aylik = df_g.groupby("Ay-Yıl")[["Net Prim", "Aktürk Sigorta Kazancı"]].sum().reset_index()
                            grup_aylik.rename(columns={"Net Prim": "Net Üretim"}, inplace=True)
                            st.bar_chart(grup_aylik.set_index("Ay-Yıl"), color=["#2563EB", "#10B981"])

    elif menu == "📅 Yenileme Takvimi":
        st.header("📅 Yenileme Takibi")
        df_pol = get_data("Policeler")
        t_col1, t_col2 = st.columns(2)
        takvim_basla = t_col1.date_input("Takvim Başlangıç", datetime.today())
        takvim_bitis = t_col2.date_input("Takvim Bitiş", datetime.today() + timedelta(days=30))
        st.divider()

        if not df_pol.empty and "Bitiş Tarihi" in df_pol.columns:
            df_pol['Bit_Obj'] = pd.to_datetime(df_pol['Bitiş Tarihi'], dayfirst=True, errors='coerce')
            takvim_temel = df_pol[(df_pol['Bit_Obj'].dt.date >= takvim_basla) & (df_pol['Bit_Obj'].dt.date <= takvim_bitis)].copy()
            takvim_temel = takvim_temel[~takvim_temel["Sigorta Şirketi"].str.contains("İPTAL", na=False)]
            takvim_temel = takvim_temel[takvim_temel["Başlangıç Tarihi"] != takvim_temel["Bitiş Tarihi"]]

            iptal_df = df_pol[df_pol["Sigorta Şirketi"].str.contains("İPTAL-SATIŞ", na=False)]
            iptal_plakalar = iptal_df[iptal_df["Plaka"] != ""]["Plaka"].unique()
            iptal_musteriler = iptal_df[iptal_df["Plaka"] == ""]["Müşteri Adı Soyadı"].unique()
            
            takvim_temel = takvim_temel[~takvim_temel["Plaka"].isin(iptal_plakalar)]
            takvim_temel = takvim_temel[~((takvim_temel["Plaka"] == "") & (takvim_temel["Müşteri Adı Soyadı"].isin(iptal_musteriler)))]
            takvim_temel["Kalan Gün"] = (takvim_temel['Bit_Obj'].dt.normalize() - pd.Timestamp.today().normalize()).dt.days
            takvim = takvim_temel.sort_values("Kalan Gün")
            
            if not takvim.empty:
                def uyari(gun):
                    if gun < 0: return "🚨 SÜRESİ DOLDU"
                    if gun <= 10: return "🔴 ACİL YENİLEME"
                    if gun <= 30: return "🟡 YAKLAŞIYOR"
                    return "🟢 SÜRESİ VAR"
                takvim["DURUM"] = takvim["Kalan Gün"].apply(uyari)
                gosterim = ["DURUM", "Kalan Gün", "Bitiş Tarihi", "Müşteri Adı Soyadı", "Plaka", "Sigorta Türü", "Sigorta Şirketi", "Telefon / E-mail"]
                if "PDF Linki" in takvim.columns: gosterim.append("PDF Linki")
                st.dataframe(df_gorsel_yap(takvim, [])[gosterim], column_config=STIL_AYARLARI, use_container_width=True)
                excel_indir(takvim[gosterim], "Bu Takvimi Excel İndir", "Ozel_Yenileme_Takvimi")
            else: st.info("Bu tarihler arasında süresi dolacak poliçe bulunmuyor.")

    elif menu == "🔎 Akıllı Arama":
        st.header("🔎 Akıllı & Bağlamlı Arama")
        df_pol = get_data("Policeler")
        if not df_pol.empty:
            ara = st.text_input("🔍 Müşteri Adı veya Plaka Yazın (Örn: 06EJA):")
            if ara:
                ara_temiz = temiz_isim(ara)
                mask = df_pol['Müşteri Adı Soyadı'].str.contains(ara_temiz, na=False) | df_pol['Plaka'].str.contains(ara_temiz, na=False) | df_pol['Poliçe No'].str.contains(ara_temiz, na=False)
                ilk_sonuc = df_pol[mask].copy()
                
                if not ilk_sonuc.empty:
                    df_pol['Kök_Poliçe'] = df_pol['Poliçe No'].apply(get_kok_police)
                    eslesen_kokler = [k for k in df_pol.loc[mask, 'Kök_Poliçe'].dropna().unique().tolist() if str(k).strip() != ""]
                    sonuc = df_pol[df_pol['Kök_Poliçe'].isin(eslesen_kokler) | mask].copy() if eslesen_kokler else ilk_sonuc.copy()
                    
                    tab1, tab2 = st.tabs(["🔗 Bağlamlı Görünüm", "📋 Klasik Liste Görünümü"])
                    with tab1:
                        sonuc['Baglam_Key'] = sonuc.apply(lambda x: f"Kök Poliçe No: {x['Kök_Poliçe']} | Ürün: {str(x['Sigorta Türü']).strip()}" if str(x['Kök_Poliçe']).strip() != "" else f"Tekil Kayıt (No Yok) | {str(x['Sigorta Türü']).strip()}_{x.name}", axis=1)
                        for key, grup in sonuc.groupby('Baglam_Key'):
                            net_toplam = grup["Net Prim"].apply(sayiya_cevir).sum(); brut_toplam = grup["Brüt Prim"].apply(sayiya_cevir).sum(); kom_toplam = grup["Şirket Komisyonu"].apply(sayiya_cevir).sum()
                            durum = "🟢 AKTİF" if brut_toplam > 0 else ("🔴 TAMAMEN İPTAL" if len(grup) > 1 else "⚠️ DİKKAT")
                            with st.expander(f"📁 {grup.iloc[0]['Müşteri Adı Soyadı']} -> {key.split('_')[0] if 'Tekil Kayıt' in key else key} | Kalan Brüt: {para_format(brut_toplam)} | {durum}"):
                                st.markdown(f"**💰 Poliçenin Güncel Net Primi:** {para_format(net_toplam)} | **Kalan Net Komisyon:** {para_format(kom_toplam)}")
                                st.dataframe(df_gorsel_yap(grup, ["Net Prim", "Brüt Prim", "Şirket Komisyonu"])[["Tanzim Tarihi", "Sigorta Şirketi", "Plaka", "Net Prim", "Brüt Prim", "Şirket Komisyonu", "PDF Linki"]], column_config=STIL_AYARLARI, use_container_width=True)
                    with tab2:
                        df_ui_arama = df_gorsel_yap(sonuc, ["Net Prim", "Brüt Prim", "Şirket Komisyonu"])
                        st.dataframe(df_ui_arama[[c for c in df_ui_arama.columns if c not in ["Sheet_Row", "Baglam_Key", "Kök_Poliçe"]]], column_config=STIL_AYARLARI, use_container_width=True)
                        excel_indir(sonuc, "Arama Sonuçlarını Excel İndir", f"Arama_Sonucu_{ara_temiz}")
                else: 
                    st.warning("Eşleşen tam kayıt bulunamadı.")
                    tum = [str(x) for x in df_pol['Müşteri Adı Soyadı'].dropna().unique().tolist() + df_pol['Plaka'].dropna().unique().tolist() if str(x).strip() != ""]
                    if oneriler := difflib.get_close_matches(ara_temiz, tum, n=3, cutoff=0.6): st.info(f"💡 **Bunu mu demek istediniz:** {', '.join(oneriler)}")

    elif menu == "🛠️ Kayıt Onarım & Silme":
        st.header("🛠️ Sistem Kayıt Yönetimi")
        t1, t2, t3, t4, t5, t6 = st.tabs(["👤 Müşteri Bilgisi Düzelt", "🗑️ Müşteriyi Tamamen Sil", "🗑️ Poliçe Sil", "🗑️ Serbest Cari Kaydı Sil", "🚑 Rakam/Cari Onar", "🔄 İsim / Kurum Birleştir"])
        
        with t1:
            df_mus = get_data("Musteriler")
            if not df_mus.empty:
                mus_listesi = [m for m in df_mus["Musteri_Adi"].dropna().unique().tolist() if str(m).strip() != ""]
                secilen_mus = st.selectbox("Düzeltilecek Müşteriyi Seçin:", ["Seçiniz..."] + sorted(mus_listesi))
                if secilen_mus != "Seçiniz...":
                    mus_bilgi = df_mus[df_mus["Musteri_Adi"] == secilen_mus].iloc[0]
                    with st.form("mus_duzelt"):
                        yeni_isim = st.text_input("Müşteri Adı Soyadı", value=secilen_mus)
                        yeni_tc = st.text_input("TC / VKN", value=str(mus_bilgi.get("TC_VKN", "")))
                        yeni_tel = st.text_input("Telefon", value=str(mus_bilgi.get("Telefon", "")))
                        if st.form_submit_button("💾 Güncelle ve Tüm Sisteme Uygula"):
                            with st.spinner("Güncelleniyor..."):
                                def _mus_guncelle():
                                    doc = client.open_by_key(SHEET_ID); y_isim_temiz = temiz_isim(yeni_isim)
                                    ws_mus = doc.worksheet("Musteriler"); m_headers = ws_mus.row_values(1); mus_satir = int(mus_bilgi["Sheet_Row"])
                                    updates = []
                                    if "Musteri_Adi" in m_headers: updates.append(gspread.Cell(row=mus_satir, col=m_headers.index("Musteri_Adi")+1, value=y_isim_temiz))
                                    if "TC_VKN" in m_headers: updates.append(gspread.Cell(row=mus_satir, col=m_headers.index("TC_VKN")+1, value=yeni_tc))
                                    if "Telefon" in m_headers: updates.append(gspread.Cell(row=mus_satir, col=m_headers.index("Telefon")+1, value=yeni_tel))
                                    if updates: ws_mus.update_cells(updates)
                                    if y_isim_temiz != secilen_mus and y_isim_temiz != "":
                                        df_pol = get_data("Policeler"); ws_pol = doc.worksheet("Policeler"); p_headers = ws_pol.row_values(1)
                                        p_updates = [gspread.Cell(row=int(row["Sheet_Row"]), col=p_headers.index("Müşteri Adı Soyadı")+1, value=y_isim_temiz) for idx, row in df_pol[df_pol["Müşteri Adı Soyadı"] == secilen_mus].iterrows() if "Müşteri Adı Soyadı" in p_headers]
                                        if p_updates: ws_pol.update_cells(p_updates)
                                        df_cari = get_data("Cari_Islemler"); ws_cari = doc.worksheet("Cari_Islemler"); c_headers = ws_cari.row_values(1)
                                        c_updates = [gspread.Cell(row=int(row["Sheet_Row"]), col=c_headers.index("Kisi_Kurum")+1, value=y_isim_temiz) for idx, row in df_cari[(df_cari["Kisi_Kurum"] == secilen_mus) & (df_cari["Islem_Turu"] == "Müşteri Carisi")].iterrows() if "Kisi_Kurum" in c_headers]
                                        if c_updates: ws_cari.update_cells(c_updates)
                                    return y_isim_temiz
                                if api_kalkani(_mus_guncelle): st.success("Başarılı!"); st.cache_data.clear(); st.rerun()

        with t2:
            st.warning("⚠️ DİKKAT: Bu işlem müşteriyi, müşteriye ait tüm cari/finans geçmişini ve Google Sheet'teki kayıtlarını KALICI olarak siler.")
            df_mus_sil = get_data("Musteriler")
            if not df_mus_sil.empty:
                sil_mus_listesi = [m for m in df_mus_sil["Musteri_Adi"].dropna().unique().tolist() if str(m).strip() != ""]
                silinecek_mus = st.selectbox("Tamamen Silinecek Müşteriyi Seçin:", ["Seçiniz..."] + sorted(sil_mus_listesi))
                if silinecek_mus != "Seçiniz...":
                    if st.button(f"🚨 {silinecek_mus} İsimli Müşteriyi SİL", type="primary"):
                        with st.spinner(f"{silinecek_mus} siliniyor..."):
                            def _mus_tamamen_sil():
                                doc = client.open_by_key(SHEET_ID)
                                ws_mus = doc.worksheet("Musteriler")
                                for s in sorted(df_mus_sil[df_mus_sil["Musteri_Adi"] == silinecek_mus]["Sheet_Row"].astype(int).tolist(), reverse=True): ws_mus.delete_rows(s)
                                df_c_sil = get_data("Cari_Islemler")
                                if not df_c_sil.empty:
                                    ws_cari = doc.worksheet("Cari_Islemler")
                                    for s in sorted(df_c_sil[df_c_sil["Kisi_Kurum"] == silinecek_mus]["Sheet_Row"].astype(int).tolist(), reverse=True): ws_cari.delete_rows(s)
                                return True
                            if api_kalkani(_mus_tamamen_sil): st.success("Silindi!"); st.cache_data.clear(); st.rerun()

        with t3:
            df_pol = get_data("Policeler")
            if not df_pol.empty:
                ara_pol_t = temiz_isim(st.text_input("Silmek istediğiniz poliçeyi arayın (Plaka, Poliçe No, Müşteri):").upper())
                if ara_pol_t:
                    sonuc = df_pol[df_pol['Müşteri Adı Soyadı'].str.contains(ara_pol_t, na=False) | df_pol['Plaka'].str.contains(ara_pol_t, na=False) | df_pol['Poliçe No'].str.contains(ara_pol_t, na=False)]
                    if not sonuc.empty:
                        silinecek_secim = st.selectbox("Silinecek Poliçeyi Seçin:", sonuc.apply(lambda x: f"{x['Tanzim Tarihi']} | {x['Müşteri Adı Soyadı']} | {x['Plaka']} | Brüt: {x['Brüt Prim']} (Satır:{x['Sheet_Row']})", axis=1).tolist())
                        if st.button("🚨 Seçili Poliçeyi TAMAMEN SİL", type="primary"):
                            with st.spinner("Poliçe ve Cari Kayıtlar siliniyor..."):
                                def _police_sil():
                                    satir_no = int(re.search(r'\(Satır:(\d+)\)', silinecek_secim).group(1)); s_row = sonuc[sonuc["Sheet_Row"] == satir_no].iloc[0]
                                    drive_pdf_sil(s_row.get("PDF Linki", "Yok")); doc = client.open_by_key(SHEET_ID); doc.worksheet("Policeler").delete_rows(satir_no)
                                    aciklama_koku = f"{s_row['Sigorta Şirketi']} - {s_row['Sigorta Türü']} - Plaka: {s_row['Plaka']}".replace(" (İPTAL-SATIŞ)","").replace(" (İPTAL-ZEYL)","")
                                    df_cari = get_data("Cari_Islemler")
                                    if not df_cari.empty:
                                        ws_cari = doc.worksheet("Cari_Islemler")
                                        for c_satir in sorted(df_cari[df_cari["Islem_Detayi"].str.contains(aciklama_koku, regex=False, na=False)]["Sheet_Row"].astype(int).tolist(), reverse=True): ws_cari.delete_rows(c_satir)
                                    return True
                                if api_kalkani(_police_sil): st.success("Silindi!"); st.cache_data.clear(); st.rerun()

        with t4:
            df_cari = get_data("Cari_Islemler")
            if not df_cari.empty:
                ara_c_t = temiz_isim(st.text_input("Silinecek Cari Kaydını Arayın (Müşteri, Tutar, Açıklama vb.):").upper())
                if ara_c_t:
                    sonuc_c = df_cari[df_cari['Kisi_Kurum'].str.contains(ara_c_t, na=False) | df_cari['Islem_Detayi'].str.contains(ara_c_t, na=False) | df_cari['Borc'].astype(str).str.contains(ara_c_t, na=False) | df_cari['Alacak'].astype(str).str.contains(ara_c_t, na=False)]
                    if not sonuc_c.empty:
                        sil_c_sec = st.selectbox("Silinecek İşlemi Seçin:", sonuc_c.apply(lambda x: f"{x['Tarih']} | {x['Kisi_Kurum']} | {x['Islem_Detayi']} | Borç: {x['Borc']} / Alacak: {x['Alacak']} (Satır:{x['Sheet_Row']})", axis=1).tolist())
                        if st.button("🚨 Bu Cari Kaydını SİL", type="primary"):
                            with st.spinner("Siliniyor..."):
                                def _cari_sil(): client.open_by_key(SHEET_ID).worksheet("Cari_Islemler").delete_rows(int(re.search(r'\(Satır:(\d+)\)', sil_c_sec).group(1))); return True
                                if api_kalkani(_cari_sil): st.success("Silindi!"); st.cache_data.clear(); st.rerun()

        with t5:
            df_pol = get_data("Policeler")
            if not df_pol.empty:
                ara_hata_t = temiz_isim(st.text_input("Düzeltilecek Müşterinin Adını veya Plakasını Yazın:"))
                hatali_df = df_pol[df_pol['Müşteri Adı Soyadı'].str.contains(ara_hata_t, na=False) | df_pol['Plaka'].str.contains(ara_hata_t, na=False)].copy() if ara_hata_t else df_pol.assign(Brüt_Siralama=df_pol["Brüt Prim"].apply(sayiya_cevir)).sort_values("Brüt_Siralama", ascending=False).head(10).copy()
                
                if not hatali_df.empty:
                    df_to_edit_hata = hatali_df[["Müşteri Adı Soyadı", "Plaka", "Sigorta Türü", "Acente", "Net Prim", "Brüt Prim", "Şirket Komisyonu"]].copy()
                    for c in ["Net Prim", "Brüt Prim", "Şirket Komisyonu"]: df_to_edit_hata[c] = df_to_edit_hata[c].apply(editor_icin_hazirla)
                    edited_hata_df = st.data_editor(df_to_edit_hata, key="hata_motoru", disabled=["Müşteri Adı Soyadı", "Plaka", "Sigorta Türü", "Acente"], use_container_width=True)
                    
                    if st.button("🚀 Düzeltmeleri Kaydet ve CARİLERE YANSIT", type="primary"):
                        with st.spinner("Onarılıyor..."):
                            def _hata_onar():
                                doc = client.open_by_key(SHEET_ID); ws_pol, ws_cari = doc.worksheet("Policeler"), doc.worksheet("Cari_Islemler")
                                df_cari = get_data("Cari_Islemler"); df_acenteler = get_data("Ayarlar_Acenteler")
                                headers_pol, headers_cari = ws_pol.row_values(1), ws_cari.row_values(1)
                                acente_oranlari = dict(zip(df_acenteler['Acente_Adi'], df_acenteler['Tali_Oran'])) if not df_acenteler.empty else {}
                                cells_to_update_pol, cells_to_update_cari = [], []
                                
                                for idx in df_to_edit_hata.index:
                                    old_row, new_row = df_to_edit_hata.loc[idx], edited_hata_df.loc[idx]
                                    eski_net, yeni_net = float(sayiya_cevir(old_row["Net Prim"])), float(sayiya_cevir(new_row["Net Prim"]))
                                    eski_brut, yeni_brut = float(sayiya_cevir(old_row["Brüt Prim"])), float(sayiya_cevir(new_row["Brüt Prim"]))
                                    eski_kom, yeni_kom = float(sayiya_cevir(old_row["Şirket Komisyonu"])), float(sayiya_cevir(new_row["Şirket Komisyonu"]))
                                    
                                    if eski_net != yeni_net or eski_brut != yeni_brut or eski_kom != yeni_kom:
                                        sheet_row_pol = int(hatali_df.loc[idx, "Sheet_Row"])
                                        if eski_net != yeni_net: cells_to_update_pol.append(gspread.Cell(row=sheet_row_pol, col=headers_pol.index("Net Prim")+1, value=yeni_net))
                                        if eski_brut != yeni_brut: cells_to_update_pol.append(gspread.Cell(row=sheet_row_pol, col=headers_pol.index("Brüt Prim")+1, value=yeni_brut))
                                        if eski_kom != yeni_kom: cells_to_update_pol.append(gspread.Cell(row=sheet_row_pol, col=headers_pol.index("Şirket Komisyonu")+1, value=yeni_kom))
                                        
                                        mus, plk, urn, acn, sir = hatali_df.loc[idx, "Müşteri Adı Soyadı"], hatali_df.loc[idx, "Plaka"], hatali_df.loc[idx, "Sigorta Türü"], hatali_df.loc[idx, "Acente"], hatali_df.loc[idx, "Sigorta Şirketi"]
                                        aciklama_koku = f"{sir.replace(' (İPTAL-SATIŞ)','').replace(' (İPTAL-ZEYL)','')} - {urn} - Plaka: {plk}"
                                        
                                        if not df_cari.empty:
                                            for c_idx, c_row in df_cari[(df_cari["Kisi_Kurum"] == mus) & (df_cari["Islem_Turu"] == "Müşteri Carisi") & (df_cari["Islem_Detayi"].str.contains(aciklama_koku, regex=False, na=False))].iterrows():
                                                eski_borc, eski_alacak = float(sayiya_cevir(c_row["Borc"])), float(sayiya_cevir(c_row["Alacak"]))
                                                if eski_borc > 0: cells_to_update_cari.append(gspread.Cell(row=int(c_row["Sheet_Row"]), col=headers_cari.index("Borc")+1, value=abs(yeni_brut)))
                                                elif eski_alacak > 0: cells_to_update_cari.append(gspread.Cell(row=int(c_row["Sheet_Row"]), col=headers_cari.index("Alacak")+1, value=abs(yeni_brut)))
                                                    
                                            if acn != "AKTÜRK SİGORTA (MERKEZ)":
                                                yeni_kazanc = yeni_kom * (float(sayiya_cevir(acente_oranlari.get(acn, 0.0))) / 100 if float(sayiya_cevir(acente_oranlari.get(acn, 0.0))) > 1 else float(sayiya_cevir(acente_oranlari.get(acn, 0.0))))
                                                for c_idx, c_row in df_cari[(df_cari["Kisi_Kurum"] == acn) & (df_cari["Islem_Turu"] == "Tali Acente Carisi") & (df_cari["Islem_Detayi"].str.contains(aciklama_koku, regex=False, na=False))].iterrows():
                                                    eski_borc, eski_alacak = float(sayiya_cevir(c_row["Borc"])), float(sayiya_cevir(c_row["Alacak"]))
                                                    if eski_borc > 0: cells_to_update_cari.append(gspread.Cell(row=int(c_row["Sheet_Row"]), col=headers_cari.index("Borc")+1, value=abs(yeni_kazanc)))
                                                    elif eski_alacak > 0: cells_to_update_cari.append(gspread.Cell(row=int(c_row["Sheet_Row"]), col=headers_cari.index("Alacak")+1, value=abs(yeni_kazanc)))
                                            else:
                                                for c_idx, c_row in df_cari[(df_cari["Islem_Turu"] == "Sigorta Şirketi Carisi") & (df_cari["Islem_Detayi"].str.contains(aciklama_koku, regex=False, na=False))].iterrows():
                                                    eski_borc, eski_alacak = float(sayiya_cevir(c_row["Borc"])), float(sayiya_cevir(c_row["Alacak"]))
                                                    if eski_borc > 0: cells_to_update_cari.append(gspread.Cell(row=int(c_row["Sheet_Row"]), col=headers_cari.index("Borc")+1, value=abs(yeni_kom)))
                                                    elif eski_alacak > 0: cells_to_update_cari.append(gspread.Cell(row=int(c_row["Sheet_Row"]), col=headers_cari.index("Alacak")+1, value=abs(yeni_kom)))

                                if cells_to_update_pol: ws_pol.update_cells(cells_to_update_pol, value_input_option='USER_ENTERED')
                                if cells_to_update_cari: ws_cari.update_cells(cells_to_update_cari, value_input_option='USER_ENTERED')
                                return bool(cells_to_update_pol or cells_to_update_cari)
                            if api_kalkani(_hata_onar): st.success("Onarıldı!"); st.cache_data.clear(); st.rerun()

        with t6:
            st.markdown("### 🔄 Toplu İsim Değiştirme ve Cari Birleştirme")
            df_pol_t6 = get_data("Policeler"); df_cari_t6 = get_data("Cari_Islemler")
            eski_isimler = set()
            if not df_pol_t6.empty: eski_isimler.update(df_pol_t6["Sigorta Şirketi"].str.replace(r' \(İPTAL-.*?\)', '', regex=True).str.strip().dropna().tolist()); eski_isimler.update(df_pol_t6["Acente"].dropna().tolist())
            if not df_cari_t6.empty: eski_isimler.update(df_cari_t6["Kisi_Kurum"].dropna().tolist())
            
            c_m1, c_m2 = st.columns(2)
            eski_secim = c_m1.selectbox("❌ Eski / Hatalı İsim (Silinecek):", ["Seçiniz..."] + sorted([x for x in eski_isimler if str(x).strip() != ""]))
            yeni_secim = c_m2.text_input("✅ Yeni / Doğru İsim (Birleşecek):", value="", help="Doğa Sigorta, Hepiyi Sigorta vb.")
            
            if st.button("🚀 Tüm Sistemde Değiştir ve Hesapları Birleştir", type="primary"):
                if eski_secim != "Seçiniz..." and yeni_secim.strip() != "":
                    with st.spinner(f"Güncelleniyor..."):
                        def _toplu_degistir():
                            doc = client.open_by_key(SHEET_ID); yeni_isim_temiz = temiz_isim(yeni_secim)
                            
                            if not df_cari_t6.empty:
                                ws_cari = doc.worksheet("Cari_Islemler"); c_headers = ws_cari.row_values(1)
                                if "Kisi_Kurum" in c_headers:
                                    c_updates = [gspread.Cell(row=int(row["Sheet_Row"]), col=c_headers.index("Kisi_Kurum") + 1, value=yeni_isim_temiz) for idx, row in df_cari_t6[df_cari_t6["Kisi_Kurum"] == eski_secim].iterrows()]
                                    if c_updates: ws_cari.update_cells(c_updates)
                            
                            if not df_pol_t6.empty:
                                ws_pol = doc.worksheet("Policeler"); p_headers = ws_pol.row_values(1); p_updates = []
                                if "Acente" in p_headers: p_updates.extend([gspread.Cell(row=int(row["Sheet_Row"]), col=p_headers.index("Acente") + 1, value=yeni_isim_temiz) for idx, row in df_pol_t6[df_pol_t6["Acente"] == eski_secim].iterrows()])
                                if "Sigorta Şirketi" in p_headers:
                                    s_col_idx = p_headers.index("Sigorta Şirketi") + 1
                                    for idx, row in df_pol_t6.iterrows():
                                        sir_val = str(row["Sigorta Şirketi"])
                                        if sir_val.replace(" (İPTAL-SATIŞ)", "").replace(" (İPTAL-ZEYL)", "").strip() == eski_secim: p_updates.append(gspread.Cell(row=int(row["Sheet_Row"]), col=s_col_idx, value=sir_val.replace(eski_secim, yeni_isim_temiz)))
                                if p_updates: ws_pol.update_cells(p_updates)
                            return True
                        if api_kalkani(_toplu_degistir): st.success("🎉 Başarıyla birleştirildi!"); st.cache_data.clear(); time.sleep(1); st.rerun()

    elif menu == "⚙️ Sistem Ayarları":
        st.header("⚙️ Genel Ayarlar ve Sabitler")
        t1, t2, t3, t4, t5 = st.tabs(["📊 Ürün Oranları", "🏢 Tali Acenteler", "🏢 Sigorta Şirketleri", "🔄 Veri Senk.", "📇 Hesap Planı (YENİ)"])
        
        with t1:
            df_urun = get_data("Ayarlar_Urunler")
            if df_urun.empty: df_urun = pd.DataFrame(columns=["Urun_Adi", "Komisyon_Orani"])
            edited_urun = st.data_editor(df_urun.drop(columns=["Sheet_Row"], errors='ignore'), num_rows="dynamic", use_container_width=True)
            if st.button("💾 Ürünleri Kaydet", key="btn_urun"):
                with st.spinner("Kaydediliyor..."):
                    def _urun_kaydet():
                        ws_urun = client.open_by_key(SHEET_ID).worksheet("Ayarlar_Urunler"); ws_urun.clear(); e_urun = edited_urun.fillna("")
                        data_to_save = [e_urun.columns.values.tolist()]
                        for r in e_urun.values.tolist(): data_to_save.append([float(sayiya_cevir(v)) if isinstance(v, (int, float, str)) and sayiya_cevir(v) != 0 else v for v in r])
                        ws_urun.append_rows(data_to_save, value_input_option='USER_ENTERED'); return True
                    if api_kalkani(_urun_kaydet): st.cache_data.clear(); st.rerun()
        with t2:
            df_acente = get_data("Ayarlar_Acenteler")
            if df_acente.empty: df_acente = pd.DataFrame(columns=["Acente_Adi", "Tali_Oran", "Hesap_Kodu"])
            edited_acente = st.data_editor(df_acente.drop(columns=["Sheet_Row"], errors='ignore'), num_rows="dynamic", use_container_width=True)
            if st.button("💾 Acenteleri Kaydet", key="btn_acente"):
                with st.spinner("Kaydediliyor..."):
                    def _acente_kaydet():
                        ws_acente = client.open_by_key(SHEET_ID).worksheet("Ayarlar_Acenteler"); ws_acente.clear(); e_acente = edited_acente.fillna("")
                        data_to_save = [e_acente.columns.values.tolist()]
                        for r in e_acente.values.tolist(): r[0] = temiz_isim(r[0]); data_to_save.append([float(sayiya_cevir(v)) if isinstance(v, (int, float, str)) and sayiya_cevir(v) != 0 and str(v).replace('.','',1).isdigit() else v for v in r])
                        ws_acente.append_rows(data_to_save, value_input_option='USER_ENTERED'); return True
                    if api_kalkani(_acente_kaydet): st.cache_data.clear(); st.rerun()
        with t3:
            df_sirket = get_data("Ayarlar_Sirketler")
            if df_sirket.empty: df_sirket = pd.DataFrame(columns=["Sirket_Adi", "Hesap_Kodu"])
            edited_sirket = st.data_editor(df_sirket.drop(columns=["Sheet_Row"], errors='ignore'), num_rows="dynamic", use_container_width=True)
            if st.button("💾 Şirketleri Kaydet", key="btn_sirket"):
                with st.spinner("Kaydediliyor..."):
                    def _sirket_kaydet():
                        ws_sirket = client.open_by_key(SHEET_ID).worksheet("Ayarlar_Sirketler"); ws_sirket.clear(); e_sirket = edited_sirket.fillna("")
                        data_to_save = [e_sirket.columns.values.tolist()]; 
                        for r in e_sirket.values.tolist(): r[0] = temiz_isim(r[0]); data_to_save.append(r)
                        ws_sirket.append_rows(data_to_save, value_input_option='USER_ENTERED'); return True
                    if api_kalkani(_sirket_kaydet): st.cache_data.clear(); st.rerun()
        with t4:
            st.warning("Geçmiş poliçeleri Cari İşlemler tablosuna işler.")
            if st.button("🚀 Senkronize Et"):
                with st.spinner("Taranıyor..."):
                    def _senk_et():
                        df_pol = get_data("Policeler"); df_cari = get_data("Cari_Islemler")
                        df_urun = get_data("Ayarlar_Urunler"); df_acente = get_data("Ayarlar_Acenteler"); df_mus = get_data("Musteriler")
                        urun_oranlari = dict(zip(df_urun['Urun_Adi'], df_urun['Komisyon_Orani'])) if not df_urun.empty else {}
                        acente_oranlari = dict(zip(df_acente['Acente_Adi'], df_acente['Tali_Oran'])) if not df_acente.empty else {}
                        
                        doc = client.open_by_key(SHEET_ID); ws_cari = doc.worksheet("Cari_Islemler"); c_headers = ws_cari.row_values(1)
                        if "Fiş No" not in c_headers: ws_cari.update_cell(1, len(c_headers)+1, "Fiş No"); c_headers.append("Fiş No")
                        
                        mevcut_cariler = df_cari["Islem_Detayi"].tolist() if not df_cari.empty and "Islem_Detayi" in df_cari.columns else []
                        mevcut_musteriler = [temiz_isim(m) for m in (df_mus["Musteri_Adi"].tolist() if not df_mus.empty and "Musteri_Adi" in df_mus.columns else [])]
                        yeni_satirlar_cari = []; yeni_satirlar_mus = []
                        
                        def c_satir_hazirla(tarih, tur, kurum, detay, borc, alacak, odeme, taksit, fis):
                            satir = [""] * len(c_headers)
                            mapping = {"Tarih": tarih, "Islem_Turu": tur, "Kisi_Kurum": kurum, "Islem_Detayi": detay, "Borc": borc, "Alacak": alacak, "Odeme_Tipi": odeme, "Taksit": taksit, "Fiş No": fis}
                            for key, val in mapping.items():
                                if key in c_headers: satir[c_headers.index(key)] = val
                            return satir
                            
                        for index, row in df_pol.iterrows():
                            mus = temiz_isim(str(row.get("Müşteri Adı Soyadı", ""))); sir = temiz_isim(str(row.get("Sigorta Şirketi", ""))); plk = str(row.get("Plaka", ""))
                            if mus == "" and sir == "" and plk == "": continue 
                            
                            y_fis = fis_no_uret("SNK"); tc = str(row.get("TC / VKN", "")); ilet = str(row.get("Telefon / E-mail", "")); urn = str(row.get("Sigorta Türü", ""))
                            net = float(sayiya_cevir(row.get("Net Prim", 0))); brut = float(sayiya_cevir(row.get("Brüt Prim", 0))); kom = float(sayiya_cevir(row.get("Şirket Komisyonu", 0))); acn = temiz_isim(str(row.get("Acente", "AKTÜRK SİGORTA (MERKEZ)")))
                            islem_tarihi = tarih_formatla(row.get("Tanzim Tarihi", ""))
                            islem_notu = "SATIŞ/TAM İPTAL İADESİ - " if "İPTAL-SATIŞ" in sir else ("KISMİ İPTAL/ZEYL İADESİ - " if "İPTAL-ZEYL" in sir else ("İPTAL/İADE - " if net < 0 or brut < 0 else ""))
                            aciklama = f"{islem_notu}{sir.replace(' (İPTAL-SATIŞ)','').replace(' (İPTAL-ZEYL)','')} - {urn} - Plaka: {plk}"

                            if aciklama not in mevcut_cariler and f"Acente Payı Kesintisi - {aciklama}" not in mevcut_cariler and f"Şirket Komisyonu Hakediş - {aciklama}" not in mevcut_cariler:
                                sirket_komisyonu = kom if kom != 0.0 else net * (float(sayiya_cevir(urun_oranlari.get(urn, 0.0))) / 100 if float(sayiya_cevir(urun_oranlari.get(urn, 0.0))) > 1 else float(sayiya_cevir(urun_oranlari.get(urn, 0.0))))
                                akturk_kazanci = float(sirket_komisyonu * (float(sayiya_cevir(acente_oranlari.get(acn, 0.0))) / 100 if float(sayiya_cevir(acente_oranlari.get(acn, 0.0))) > 1 else float(sayiya_cevir(acente_oranlari.get(acn, 0.0)))))

                                yeni_satirlar_cari.append(c_satir_hazirla(islem_tarihi, "Müşteri Carisi", mus, aciklama, brut, 0.0, "Aktarılmış Kayıt", 1, y_fis))
                                if acn != "AKTÜRK SİGORTA (MERKEZ)": yeni_satirlar_cari.append(c_satir_hazirla(islem_tarihi, "Tali Acente Carisi", acn, f"Acente Payı Kesintisi - {aciklama}", akturk_kazanci, 0.0, "Aktürk Sigorta Kazancı", 1, y_fis))
                                else: yeni_satirlar_cari.append(c_satir_hazirla(islem_tarihi, "Sigorta Şirketi Carisi", sir.replace(' (İPTAL-SATIŞ)','').replace(' (İPTAL-ZEYL)',''), f"Şirket Komisyonu Hakediş - {aciklama}", sirket_komisyonu, 0.0, "Aktarılmış Kayıt", 1, y_fis))
                                mevcut_cariler.append(aciklama)

                            if mus and mus not in mevcut_musteriler: yeni_satirlar_mus.append([mus, tc, ilet, brut]); mevcut_musteriler.append(mus) 

                        if yeni_satirlar_cari: ws_cari.append_rows(yeni_satirlar_cari, value_input_option='USER_ENTERED') 
                        if yeni_satirlar_mus: doc.worksheet("Musteriler").append_rows(yeni_satirlar_mus, value_input_option='USER_ENTERED')
                        return bool(yeni_satirlar_cari or yeni_satirlar_mus)
                        
                    if api_kalkani(_senk_et): st.success("Tamamlandı."); st.cache_data.clear()
        with t5:
            st.markdown("### 📇 Otomatik Hesap Planı (Kalıcı Kod Sabitleme)")
            if st.button("🚀 Hesap Planını Oluştur / Kodları Sabitle", type="primary"):
                with st.spinner("Google Sheets taranıyor..."):
                    def _hesap_plani_olustur():
                        doc = client.open_by_key(SHEET_ID)
                        ws_mus = doc.worksheet("Musteriler"); mus_data = ws_mus.get_all_values()
                        if mus_data:
                            headers = mus_data[0]
                            if "Hesap_Kodu" not in headers: ws_mus.update_cell(1, len(headers)+1, "Hesap_Kodu"); headers.append("Hesap_Kodu")
                            hk_idx = headers.index("Hesap_Kodu"); mevcut_kodlar = [row[hk_idx] for row in mus_data[1:] if len(row) > hk_idx and row[hk_idx].startswith("120.")]
                            max_kod = max([int(k.split(".")[-1]) for k in mevcut_kodlar] + [0]); updates = []
                            for i, row in enumerate(mus_data[1:]):
                                if len(row) <= hk_idx or not str(row[hk_idx]).startswith("120."): max_kod += 1; updates.append(gspread.Cell(row=i+2, col=hk_idx+1, value=f"120.{str(max_kod).zfill(3)}"))
                            if updates: ws_mus.update_cells(updates)
                        
                        ws_acn = doc.worksheet("Ayarlar_Acenteler"); acn_data = ws_acn.get_all_values()
                        if acn_data:
                            headers = acn_data[0]
                            if "Hesap_Kodu" not in headers: ws_acn.update_cell(1, len(headers)+1, "Hesap_Kodu"); headers.append("Hesap_Kodu")
                            hk_idx = headers.index("Hesap_Kodu"); mevcut_kodlar = [row[hk_idx] for row in acn_data[1:] if len(row) > hk_idx and row[hk_idx].startswith("320.T.")]
                            max_kod = max([int(k.split(".")[-1]) for k in mevcut_kodlar] + [0]); updates = []
                            for i, row in enumerate(acn_data[1:]):
                                if len(row) <= hk_idx or not str(row[hk_idx]).startswith("320.T."): max_kod += 1; updates.append(gspread.Cell(row=i+2, col=hk_idx+1, value=f"320.T.{str(max_kod).zfill(3)}"))
                            if updates: ws_acn.update_cells(updates)

                        try:
                            ws_sir = doc.worksheet("Ayarlar_Sirketler"); sir_data = ws_sir.get_all_values()
                            if sir_data:
                                headers = sir_data[0]
                                if "Hesap_Kodu" not in headers: ws_sir.update_cell(1, len(headers)+1, "Hesap_Kodu"); headers.append("Hesap_Kodu")
                                hk_idx = headers.index("Hesap_Kodu"); mevcut_kodlar = [row[hk_idx] for row in sir_data[1:] if len(row) > hk_idx and row[hk_idx].startswith("320.S.")]
                                max_kod = max([int(k.split(".")[-1]) for k in mevcut_kodlar] + [0]); updates = []
                                for i, row in enumerate(sir_data[1:]):
                                    if len(row) <= hk_idx or not str(row[hk_idx]).startswith("320.S."): max_kod += 1; updates.append(gspread.Cell(row=i+2, col=hk_idx+1, value=f"320.S.{str(max_kod).zfill(3)}"))
                                if updates: ws_sir.update_cells(updates)
                        except: pass
                        return True
                    if api_kalkani(_hesap_plani_olustur): st.success("🎉 Müşteri, Şirket ve Tali Acente kodları Google Sheets'e işlendi!"); st.cache_data.clear(); st.rerun()

    elif menu == "📂 Genel Arşiv":
        st.header("📂 Tüm Poliçe Arşivi")
        df_pol = get_data("Policeler")
        if not df_pol.empty:
            df_ui_arsiv = df_gorsel_yap(df_pol, ["Net Prim", "Brüt Prim", "Şirket Komisyonu"])
            st.dataframe(df_ui_arsiv[[c for c in df_ui_arsiv.columns if c != "Sheet_Row"]], column_config=STIL_AYARLARI, use_container_width=True)
            st.divider()
            excel_indir(df_pol, "Tüm Arşivi Excel Olarak İndir", "Tum_Police_Arsivi")

    st.sidebar.divider()
    if st.sidebar.button("🚪 Güvenli Çıkış", type="secondary"): st.session_state["giris_yapildi"] = False; st.rerun()
